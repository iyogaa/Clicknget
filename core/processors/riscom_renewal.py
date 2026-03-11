from datetime import datetime
import string
from io import BytesIO
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from dateutil import parser
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import PieChart3D, BarChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    RichTextProperties,
    ParagraphProperties,
    Paragraph,
    CharacterProperties,
)
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_NUMBER
from pydantic import BaseModel, Field
import re
from ftfy import fix_encoding
from unidecode import unidecode
from thefuzz import fuzz
from openpyxl.cell import MergedCell
import io
import zipfile
from dateutil.relativedelta import relativedelta
import os
import sys
from dotenv import load_dotenv
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__)))

if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)
#hello

    

load_dotenv()



DATE_FORMAT_STRING = "%m/%d/%Y"

def unpack_tuples(cell_value):
    if isinstance(cell_value, tuple):
        return cell_value[-1]
    return cell_value

first_sheet = True
blue_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
blue_header_fill = PatternFill(
    start_color="0033CCCC", end_color="0033CCCC", fill_type="solid"
)


class Chart(BaseModel):
    title: str = ""
    data_length: int = 5
    chart_type: str
    min_col_label: int
    max_col_label: Optional[int] = None
    min_col_data: int
    max_col_data: int
    min_row_label: int
    min_row_data: int
    max_row_label: int
    max_row_data: int
    excel_column: str
    excel_row: int
    orient: Optional[str] = "col"
    dimensions: int = 2


def attach_pie_chart(worksheet, chart_obj: Chart):
    datalabel = DataLabelList()
    datalabel.showPercent = True
    # datalabel.showVal = True
    pie = PieChart3D(dLbls=datalabel)
    labels = Reference(
        worksheet,
        min_col=chart_obj.min_col_label,
        min_row=chart_obj.min_row_label,
        max_row=chart_obj.data_length + 2,
    )
    data = Reference(
        worksheet,
        min_col=chart_obj.min_col_data,
        min_row=chart_obj.min_row_data,
        max_row=chart_obj.data_length + 2,
        max_col=chart_obj.max_col_data,
    )
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    pie.title = chart_obj.title

    # pie.legend.position = None
    # pie.legend.overlay = '0'

    worksheet.add_chart(pie, f"{chart_obj.excel_column}{chart_obj.excel_row}")


def attach_bar_chart(worksheet, chart_obj: Chart, chart_first):
    if chart_first:
        start_row = 30
    else:
        start_row = 2
    axis = CharacterProperties(sz=1000)
    rot = (
        RichTextProperties(vert="vert270")
        if chart_obj.orient == "col"
        else RichTextProperties()
    )
    datalabel = DataLabelList()
    datalabel.showPercent = True
    datalabel.showVal = True
    datalabel.textProperties = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot
    )

    chart_size = chart_obj.data_length * 3
    bar_chart = (
        BarChart3D(dLbls=datalabel)
        if chart_obj.dimensions == 3
        else BarChart(dLbls=datalabel)
    )
    bar_chart.height, bar_chart.width = (
        (bar_chart.height, max(5 + chart_size, bar_chart.width))
        if chart_obj.orient == "col"
        else (max(chart_size, bar_chart.height), bar_chart.width)
    )
    bar_chart.type = chart_obj.orient
    labels = Reference(
        worksheet,
        min_col=chart_obj.min_col_label,
        min_row=chart_obj.min_row_label,
        max_row=chart_obj.data_length + start_row,
        max_col=(
            chart_obj.max_col_label
            if chart_obj.max_col_label
            else chart_obj.min_col_label
        ),
    )
    data = Reference(
        worksheet,
        min_col=chart_obj.min_col_data,
        min_row=chart_obj.min_row_data,
        max_row=chart_obj.data_length + start_row,
        max_col=chart_obj.max_col_data,
    )
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(labels)

    # for series in bar_chart.series:
    #     for point in series.points:
    #         point.dLbls.showVal = True

    bar_chart.title = chart_obj.title
    worksheet.add_chart(bar_chart, f"{chart_obj.excel_column}{chart_obj.excel_row}")


def read_exp_drivers_excel(drivers_exp_file, sheet_name=0):
    try:
        return pd.read_excel(drivers_exp_file, sheet_name=sheet_name, dtype={'LicenseNumber': str})
    except Exception as e1:
        try:
            return pd.read_excel(drivers_exp_file, sheet_name=sheet_name, engine='xlrd', dtype={'LicenseNumber': str})
        except Exception as e2:
            raise Exception("Either Driver's file or MVR file is missing or unreadable!")


def apply_formatting_new(worksheet, min_col, max_col):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(min_row=29, min_col=min_col, max_col=max_col):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 30:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row > 2 and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            cell.number_format = FORMAT_NUMBER
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def apply_formatting(worksheet, min_col, max_col, custom_format_col_list=None, custom_format_cell_list=None):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(min_col=min_col, max_col=max_col):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 2:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row > 2 and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            # if column or the cell from summary sheet is to be kept integer
            if (custom_format_col_list and not isinstance(cell,
                                                          MergedCell) and cell.column_letter in custom_format_col_list) or (
                    custom_format_cell_list and not isinstance(cell,
                                                               MergedCell) and cell.coordinate in custom_format_cell_list):
                cell.number_format = FORMAT_NUMBER
            else:
                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def add_blue_header(sheet, header_text, chart_first=False, last_column=5, start_row=None):
    if not start_row:
        sheet.insert_rows(1)
    elif start_row and start_row == 2:
        sheet.insert_rows(1)
    else:
        sheet.insert_rows(start_row - 2)

    if chart_first:
        start_row = 29
    if start_row:
        start_row = start_row - 1
    else:
        start_row = start_row if start_row else 1
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_column)

    header_cell = sheet.cell(row=start_row, column=1)
    header_cell.value = header_text
    header_cell.font = blue_header_font  # Font with smaller size
    header_cell.fill = blue_header_fill
    header_cell.alignment = Alignment(horizontal="center")


def strip_non_numeric_and_leading_zeros(input_string):
    numeric_part = re.sub(r'\D', '', str(input_string))
    return numeric_part.lstrip('0')


def match_driver_names(name_driver, name_grouped):
    set_ratio = fuzz.token_set_ratio(name_driver, name_grouped)
    sort_ratio = fuzz.token_sort_ratio(name_driver, name_grouped)
    return (set_ratio + sort_ratio) / 2


def filter_drivers_rsc(driver_row, grouped_row):
    name_driver = driver_row['Driver First Name'] + " " + driver_row['Driver Last Name']
    name_grouped = grouped_row['Driver First Name'] + " " + grouped_row['Driver Last Name']
    cdl_driver, cdl_grouped = driver_row['License Number'], grouped_row['License Number']
    dob_driver, dob_grouped = driver_row['Driver Date of Birth'], grouped_row['Driver Date of Birth']

    if strip_non_numeric_and_leading_zeros(cdl_driver) == strip_non_numeric_and_leading_zeros(cdl_grouped):
        return True

    try:
        dob_driver_parsed = parser.parse(dob_driver).date()
        dob_grouped_parsed = parser.parse(dob_grouped).date()
    except (ValueError, TypeError):
        dob_driver_parsed = None
        dob_grouped_parsed = None

    if dob_driver_parsed and dob_grouped_parsed:
        if dob_driver_parsed == dob_grouped_parsed:
            return match_driver_names(name_driver, name_grouped) >= 80
    else:
        return match_driver_names(name_driver, name_grouped) >= 80

    return False


def split_driver_name(full_name):
    SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}
    if pd.isnull(full_name) or full_name.strip() == "":
        return pd.Series(["", ""])
    full_name = full_name.strip()
    if ',' in full_name:
        # Split on the first comma
        last, first = [p.strip() for p in full_name.split(',', 1)]
        first_parts = first.split()
        # Check if the last part is a suffix
        if first_parts and first_parts[-1].upper() in SUFFIXES:
            suffix = first_parts.pop(-1)
            last = f"{last} {suffix}"
        first_name = " ".join(first_parts)
        last_name = last
    else:
        # No comma: check for suffix at the end
        parts = full_name.split()
        if len(parts) >= 3 and parts[-1].upper() in SUFFIXES:
            # e.g., "Mario Espinola II"
            last_name = f"{parts[0]} {parts[-1]}"
            first_name = " ".join(parts[1:-1])
        elif len(parts) >= 2:
            # e.g., "Smith John"
            last_name = parts[0]
            first_name = " ".join(parts[1:])
        else:
            first_name = parts[0]
            last_name = ""
    return pd.Series([first_name, last_name])


def calculate_age(dob_str):
    if dob_str:
        try:
            dob = parser.parse(dob_str).date()
            today = datetime.now().date()
            # More precise: subtract one if birthday hasn't occurred yet this year
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
            return age
        except Exception:
            return ""
    return ""


def apply_formatting_mvr_owned(
        worksheet, min_col, max_col, start_row=None, end_row=None
):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(
            min_col=min_col, max_col=max_col, min_row=start_row - 1, max_row=end_row
    ):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 2:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row >= start_row and cell.row <= end_row and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            cell.number_format = "0"
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def write_excel_sheet_mvr_owned(
        data1, workbook, sheet_name, header_text, _first_sheet, start_row=2, skip_blue_header=False
):
    # global _first_sheet
    # charts = data[1]
    start_row_real = start_row
    data = data1
    if isinstance(data, dict):
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # creating tables for the graph data
        add_blue_header(worksheet, header_text, last_column=2, start_row=start_row_real)
        # if start_row != 2:
        # start_row = start_row-1
        start_row = start_row
        start_column = 1
        for row in data.items():
            for col_index, value in enumerate(row, start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        apply_formatting_mvr_owned(
            worksheet,
            min_col=1,
            max_col=2,
            start_row=start_row_real,
            end_row=start_row_real + len(data) - 1,
        )

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].auto_size = True

    elif isinstance(data, list) and data: # Added check for empty list
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)
        start_row = start_row - 1
        start_column = 1
        last_column = 5

        header_row_flag = True
        for row in data:
            last_column = len(row)
            if header_row_flag:
                header_row_flag = False
                # Write header
                for col_index, key in enumerate(row.keys(), start=start_column):
                    worksheet.cell(
                        row=start_row,
                        column=col_index,
                        value=key
                    )
                start_row += 1
            # Write data row
            for col_index, value in enumerate(row.values(), start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        
        if not skip_blue_header:
            add_blue_header(worksheet, header_text, last_column, start_row=start_row_real)
        else:
            start_row_real -= 1  # avoid one extra grey row
        
        if data: # Only apply formatting if data exists
            apply_formatting_mvr_owned(
                worksheet,
                min_col=1,
                max_col=len(data[0]),
                start_row=start_row_real,
                end_row=start_row_real + len(data),
            )


def custom_major_minor_3_5(x, violation_dates, three_years_ago, five_years_ago):
    values = [str(value).strip().lower() for value in x]
    # Check for major violations in last 5 years
    recent_5 = [(v, d) for v, d in zip(values, violation_dates) if pd.notnull(d) and d >= five_years_ago]
    recent_5_values = [v for v, d in recent_5]
    if "major" in recent_5_values:
        return "Major"
    # Check for minor violations in last 3 years
    recent_3 = [(v, d) for v, d in zip(values, violation_dates) if pd.notnull(d) and d >= three_years_ago]
    recent_3_values = [v for v, d in recent_3]
    if "minor" in recent_3_values:
        return "Minor"
    return ""


def custom_concat(x) -> str:
    unique_values = set(x)
    non_empty_values = [value for value in unique_values if value != ""]
    res = ", ".join(map(str, non_empty_values))
    return res


def concat_without_clean_mvr(existing, to_add):
    """
    Concatenates two strings, removing 'Clean MVR' (case-insensitive) and '_x000d_\' artifacts from both.
    Adds a comma only if both are non-empty.
    If existing has 'Clean MVR' and to_add is empty, return existing as is.
    If existing has 'Clean MVR' and to_add has other values, remove 'Clean MVR' from existing.
    Also removes '_x000d_\' artifacts from both strings.
    """

    def clean_string(s):
        # Remove 'Clean MVR' (case-insensitive)
        # s = re.sub(r'(?i)\bclean mvr\b', '', str(s))
        # Remove '_x000d_\' (case-insensitive, with optional trailing whitespace/newline)
        s = re.sub(r'(?i)_x000d_\\[\s\n]*', '', s)
        # Strip leading/trailing commas and whitespace
        return s.strip(', ').strip()

    cleaned_to_add = clean_string(to_add)
    if cleaned_to_add:
        cleaned_existing = clean_string(existing)
        if cleaned_existing:
            return f"{cleaned_existing},{cleaned_to_add}"
        else:
            return cleaned_to_add
    else:
        return clean_string(existing)


def status_and_comments(row, has_heavy_vehicle):
    comments = []
    status = ""

    # Parse age
    try:
        age = int(row.get("Age", 0) or 0)
    except Exception:
        age = 0

    # Parse violation counts
    try:
        minor_count = int(row.get("Minor Count", 0) or 0)
    except Exception:
        minor_count = 0
    try:
        major_count = int(row.get("Major Count", 0) or 0)
    except Exception:
        major_count = 0

    # Parse total incidents
    try:
        total_incidents = int(row.get("Accident Count", 0) or 0)
    except Exception:
        total_incidents = 0

    # Medical status
    med_status = str(row.get("Medical Status", "")).strip().lower()
    # Medical expiration
    med_exp = row.get("Medical Expiration Date", "")
    # Date MVR Ordered
    mvr_ordered_date = row.get("Date MVR Ordered", "")
    # Restrictions
    restrictions = str(row.get("Restrictions", "")).strip()
    # Violation description
    violation_desc = row.get("Violation Description", "")

    # --- Pending conditions ---
    pending = False
    if age < 21 or age >= 70:  # not needed
        pending = True
        comments.append("Age Requirement")
    # if minor_count > 3 or major_count > 0:  # handled in post logics
    #     pending = True
    #     if violation_desc:
    #         comments.append(str(violation_desc))
    # if 'speed' in str(violation_desc).strip().lower(): #handled in post logics
    #     pending = True
    #     comments.append(violation_desc)
    if med_status == "not certified":
        pending = True
        comments.append("Medical Status Not Certified")
    if med_exp:
        try:
            med_exp_date = datetime.strptime(med_exp, "%m/%d/%Y")
            if med_exp_date < datetime.now():
                pending = True
                comments.append("Medical Expired")
        except Exception:
            pass
    if mvr_ordered_date:
        try:
            # Try to parse the date (it may be a datetime or string)
            if isinstance(mvr_ordered_date, datetime):
                mvr_date = mvr_ordered_date
            else:
                mvr_date = datetime.strptime(str(mvr_ordered_date), "%m/%d/%Y")
            if (datetime.now() - mvr_date).days > 90:
                pending = True
                comments.append("New MVR needed")
        except Exception:
            pass
    if restrictions:
        allowed_restrictions = [
            'corrective lenses',
            'corrective lens',
            'glasses/contacts',
            'with corrective lenses',
            'b - corrective lenses',
            'wear corrective lenses',
            'glasses',
            'contact lenses',
            'motorcycle'
        ]
        allowed_set = set(r.lower() for r in allowed_restrictions)

        # Split restrictions by comma, strip whitespace, and normalize case
        restriction_items = [r.strip().lower() for r in restrictions.split(',')]

        # Find any restriction not in the allowed set
        extra_restrictions = [r for r in restriction_items if r not in allowed_set]

        if extra_restrictions:
            pending = True
            comments.append(', '.join(extra_restrictions))

    if total_incidents >= 1:
        pending = True

    # New pending condition for heavy/extra heavy vehicles
    license_type = str(row.get("License Type", "")).strip().upper()
    license_number = row.get("License Number", "")
    med_exp = row.get("Medical Expiration Date", "")

    if (
            has_heavy_vehicle and
            license_type in ['A', 'B', 'C', 'D'] and
            not med_exp
    ):
        pending = True
        comments.append("Heavy/Extra Heavy Vehicle with missing Medical Expiration Date")

        # --- License Status blank ---
    license_status = str(row.get("License Status", "")).strip()
    if not license_status:
        comments.append("License status unknown")
        pending = True
    license_status = str(row.get("License Status", "")).strip().lower()
    if license_status == "invalid":
        comments.append("License Invalid")
        pending = True
    if license_status == "suspended":
        pending = True
        comments.append("License Suspended")
    if license_status == "expired":
         pending = True
         comments.append("License Expired")
    if license_status == "surrender":
         pending = True
         comments.append("License Surrender")
    if license_status == "cancelled":
        pending = True
        comments.append("License Cancelled")
    if license_status == "blocked":
        pending = True
        comments.append("License Blocked")
    if license_status in("limited"):
        comments.append("License status is LIMITED")
        pending = True
    if license_status in("temp"):
        comments.append("License status is TEMP")
        pending = True

    if pending:
        status = "Pending"
    else:
        # --- Approved conditions ---
        if (minor_count == 0 and major_count == 0 and med_status == "certified"):
            status = "Approved"
        elif (minor_count <= 3 and major_count == 0):
            status = "Approved"
        else:
            status = ""  # If none of the criteria are met

    # Add "Clean MVR" only if status is Approved and there are no other comments
    if status == "Approved" and not comments:
        comments.append("Clean MVR")

    comments_str = ", ".join([c for c in comments if c])
    return pd.Series({"Status": status, "Comments": comments_str})


def filter_and_cleanup_mvr_data(mvr_data):
    """
    Filters out rows where exp_drivers_list_status is 'deleted' and mvr_list_flag is False,
    then drops the flag/status columns from the DataFrame.
    Returns a list of dicts.
    """
    # Convert to DataFrame
    df = pd.DataFrame(mvr_data)

    # Drop the specified columns
    cols_to_drop = [
        "mvr_list_flag",
        "driver_list_flag",
        "exp_drivers_list_flag",
        "exp_drivers_list_status",
        "exp_drivers_list_comments",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count 1 Year",
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year"
    ]
    df = df.drop(columns=cols_to_drop, errors="ignore")

    # Return as list of dicts
    return df.to_dict(orient="records")


def process_exp_driver_df(df, rename_map=None):
    """
    Renames columns in the given DataFrame according to the provided mapping.
    If no mapping is provided, uses a default mapping for common driver columns.

    Args:
        df (pd.DataFrame): The DataFrame to rename columns for.
        rename_map (dict): Optional. A dictionary mapping old column names to new column names.

    Returns:
        pd.DataFrame: The DataFrame with columns renamed.
    """
    # Default mapping
    default_map = {
        "Name": "Driver Full Name",
        "LicenseNumber": "License Number",
        "DOB": "Driver Date of Birth",
        "Status": "exp_drivers_list_status",
        "LicenseClass": "License Type",
        "StateID": "License State",
        "LicenseExpDate": "License Expiration Date",
        "Comments": "exp_drivers_list_comments",
    }
    # If user provides a mapping, update the default mapping
    df["DOB"] = pd.to_datetime(df["DOB"], errors='coerce').dt.strftime("%m/%d/%Y")
    df["AddDeleteDate"] = pd.to_datetime(df["AddDeleteDate"], errors='coerce').dt.strftime("%m/%d/%Y")
    df["LicenseExpDate"] = pd.to_datetime(df["LicenseExpDate"], errors='coerce').dt.strftime("%m/%d/%Y")
    # Replace 'NaT' strings (from missing/invalid dates) with empty string
    df["DOB"] = df["DOB"].fillna("").replace("NaT", "")
    df["Comments"] = df["Comments"].fillna("")
    if rename_map:
        default_map.update(rename_map)
    return df.rename(columns=default_map)

def match_by_license(driver_row, exp_row):
    cdl_driver = driver_row['License Number']
    cdl_exp = exp_row['License Number']
    return strip_non_numeric_and_leading_zeros(cdl_driver) == strip_non_numeric_and_leading_zeros(cdl_exp)

def match_by_name_dob(driver_row, exp_row):
    name_driver = driver_row['Driver First Name'] + " " + driver_row['Driver Last Name']
    name_exp = exp_row['Driver First Name'] + " " + exp_row['Driver Last Name']
    dob_driver = driver_row['Driver Date of Birth']
    dob_exp = exp_row['Driver Date of Birth']
    try:
        dob_driver_parsed = parser.parse(dob_driver).date()
        dob_exp_parsed = parser.parse(dob_exp).date()
    except (ValueError, TypeError):
        return False
    if dob_driver_parsed == dob_exp_parsed:
        return match_driver_names(name_driver, name_exp) >= 80
    return False

def superset_drivers(drivers_df, exp_drivers_df):
    # Mark source
    drivers_df = drivers_df.copy()
    exp_drivers_df = exp_drivers_df.copy()
    drivers_df['__source'] = 'drivers'
    exp_drivers_df['__source'] = 'exp_drivers'

    # Track which exp_drivers rows are matched
    matched_exp_indices = set()
    matched_pairs = {}  # exp_index: driver_index

    unique_rows = []

    # First, add all drivers_df rows, and mark which exp_drivers they match
    for i, driver_row in drivers_df.iterrows():
        # Only consider unmatched exp_drivers rows
        unmatched_exp_drivers_df = exp_drivers_df.loc[~exp_drivers_df.index.isin(matched_exp_indices)]
        # Find matches in unmatched exp_drivers_df
        license_matches = unmatched_exp_drivers_df.apply(lambda exp_row: match_by_license(driver_row, exp_row), axis=1)
        matched_indices = license_matches[license_matches].index.tolist()
        if not matched_indices:
            # Fallback: try to match by name+dob
            fallback_matches = unmatched_exp_drivers_df.apply(lambda exp_row: match_by_name_dob(driver_row, exp_row), axis=1)
            matched_indices = fallback_matches[fallback_matches].index.tolist()
        for exp_idx in matched_indices:
            matched_exp_indices.add(exp_idx)
            matched_pairs[exp_idx] = i
        # Add driver row with both flags if matched, else only driver_list_flag
        row = driver_row.copy()
        row['driver_list_flag'] = True
        row['exp_drivers_list_flag'] = bool(matched_indices)
        if matched_indices:
            exp_row = exp_drivers_df.loc[matched_indices[0]]
            exp_driver_status = exp_row.get("exp_drivers_list_status", "").strip().lower()
            exp_driver_comments = exp_row.get("exp_drivers_list_comments", "").strip().lower()
            add_delete_date = exp_row.get("AddDeleteDate", "").strip().lower()
            row['AddDeleteDate'] = add_delete_date
            row['exp_drivers_list_status'] = exp_driver_status
            row['exp_drivers_list_comments'] = exp_driver_comments
            row['License Expiration Date'] = exp_row.get('License Expiration Date', None)
            if exp_driver_status in ["probationary driver", "probation", "excluded"]:
                # for probationary, excluded drivers, we need status from exp drivers list explicitly
                comments = row.get("Comments", "")
                if comments:
                    row["Comments"] = concat_without_clean_mvr(existing=comments, to_add=exp_driver_comments) if exp_driver_status == "approved" else f"{comments},{exp_driver_comments}"
                else:
                    row["Comments"] = exp_driver_comments
        unique_rows.append(row)

    # Now, add only unmatched exp_drivers rows
    for i, exp_row in exp_drivers_df.iterrows():
        if i not in matched_exp_indices:
            row = exp_row.copy()
            row['driver_list_flag'] = False
            row['exp_drivers_list_flag'] = True
            unique_rows.append(row)

    # Create DataFrame
    result_df = pd.DataFrame(unique_rows).drop(columns=['__source'], errors='ignore').reset_index(drop=True)
    # Ensure both columns exist and are boolean
    result_df['mvr_list_flag'] = False
    result_df['driver_list_flag'] = result_df['driver_list_flag'].astype(bool)
    result_df['exp_drivers_list_flag'] = result_df['exp_drivers_list_flag'].astype(bool)
    return result_df


def postprocess_mvr_data(mvr_data):
    """
    Modifies mvr_data in place according to expiring driver logic.
    """
    for row in mvr_data:
        exp_flag = bool(row.get("exp_drivers_list_flag"))
        drv_flag = bool(row.get("driver_list_flag"))
        mvr_flag = bool(row.get("mvr_list_flag"))
        exp_status = str(row.get("exp_drivers_list_status", "")).strip().lower()

        # Case 0: Expiring Driver List status = Deleted & also in MVR
        if exp_flag and exp_status == "deleted" and mvr_flag:
            row["Status"] = "Pending"
            row["Comments"] = "Deleted"
        # Case 1: In expiring list but not in driver list
        elif exp_flag and not drv_flag:
            row["Status"] = "Pending"
            if mvr_flag:
                row["Comments"] = "MVR Received, on expiring drivers list not on application"
            else:
                if exp_status == "excluded":
                    # for excluded we are appending in comments because we added a case
                    # for exluded where we copied comments from exp drivers list. so need to consider existing comments also
                    if row["Comments"]:
                        row["Comments"] = f"{(row['Comments'])},Excluded- not on app – no mvr received"
                    else:
                        row["Comments"] = "Excluded- not on app – no mvr received"
                else:
                    row["Comments"] = "Driver not listed on application and no MVR received"
        # Case 2: In both expiring and driver list
        elif exp_flag or drv_flag:
            if mvr_flag:
                row["MVR Received"] = "TRUE"
                # Do not add anything to Comments
            else:
                row["MVR Received"] = "FALSE"
                # Add "MVR needed" to Comments only if not already present (case-insensitive)
                comments = row.get("Comments", "")
                if "mvr needed" not in comments.lower():
                    if comments:
                        row["Comments"] = f"{comments}, MVR needed"
                    else:
                        row["Comments"] = "MVR needed"

        # New rules for exp_drivers_list_status
        exp_comments = str(row.get("exp_drivers_list_comments", "")).strip().lower()
        minor_count = int(row.get("Minor Count 3 Year", 0) or 0)  # need to consider minor violations only within 3 yrs
        minor_violation_desc = row.get("Minor Violation Description 3 Year",
                                       "")  # need to consider minor violations only woithin 3 yrs
        major_count = int(
            row.get("Major Count 1 Year", 0) or 0)  # need to consider major violation desc only within 1 yr
        major_violation_desc = row.get("Major Violation Description 1 Year",
                                       "")  # need to consider violation desc only within 1 yr
        major_violation_desc_5yr = row.get("Major Violation Description 5 Year",
                                           "")  # need to consider violation desc only within 5 yr
        major_count_5yr = int(row.get("Major Count", 0) or 0)
        comments = row.get("Comments", "")
        # NEW RULE: If any major violation, and status is Approved, and not excluded/deleted, set to Pending
        if (
                major_count > 0 and
                str(row.get("Status", "")).strip().lower() == "approved" and
                exp_status not in ["excluded", "deleted"]
        ):
            row["Status"] = "Pending"

        # If exp_status is "excluded", set Excluded and Status
        if exp_status == "excluded":
            row["Excluded"] = "Yes"
            row["Status"] = "Pending"
        if exp_status == "suspended":
            row["Status"] = "Pending"
            if "Lic Suspended" not in comments:
                row["Comments"] = f"{comments}, Lic Suspended"
        if exp_status == "invalid":
            row["Status"] = "Pending"
            if "License Invalid" not in comments:
                row["Comments"] = f"{comments}, License Invalid"

        # Pending if too many minors or any major (Approved status)
        # if exp_status in ["probationary driver", "probation"]:
        #     row["Comments"] = f"{exp_comments},{comments}"

        if exp_status == "approved" and (minor_count > 3 or major_count > 0):
            combined_desc = ", ".join(filter(None, [minor_violation_desc, major_violation_desc])).strip()
            row["Status"] = "Pending"
            if combined_desc and combined_desc.lower() not in comments.lower():
                if comments:
                    row["Comments"] = f"{comments}, {combined_desc}"
                else:
                    row["Comments"] = combined_desc
                    comments = combined_desc

        # New condition for status acc to major and minor counts
        elif exp_status == "approved" and minor_count > 0:
            # Split minor_violation_desc by comma and keep only parts with 'speed'
            speed_minor_parts = [v.strip() for v in minor_violation_desc.split(",") if "speed" in v.lower()]
            # Join the filtered minor parts
            speed_minor_desc = ", ".join(speed_minor_parts)
            # Use all major violation desc as is
            major_desc = major_violation_desc_5yr.strip()
            # Combine both (only if present)
            combined_desc = ", ".join(filter(None, [speed_minor_desc, major_desc]))
            if combined_desc and combined_desc.lower() not in comments.lower():
                if comments:
                    row["Comments"] = f"{comments}, {combined_desc}"
                else:
                    row["Comments"] = combined_desc
                    comments = combined_desc
            # Set Status to Approved if empty or not Pending
            if not row.get("Status") or row.get("Status") != "Pending":
                row["Status"] = "Approved"
        # Approved if clean (Approved status)

        # new updated conditions to pick all descriptions from mvr and comments from exp list 
        comments = row.get("Comments", "")
        # Split and normalize existing comments
        existing = set([c.strip().lower() for c in comments.split(",") if c.strip()])
        to_add = []

        # Helper to add unique items
        def add_unique_items(source):
            for item in source.split(","):
                item_clean = item.strip()
                if item_clean and item_clean.lower() not in existing:
                    to_add.append(item_clean)
                    existing.add(item_clean.lower())

        add_unique_items(major_violation_desc_5yr)
        add_unique_items(minor_violation_desc)
        add_unique_items(exp_comments)

        if to_add:
            if comments:
                row["Comments"] = f"{comments}, " + ", ".join(to_add)
            else:
                row["Comments"] = ", ".join(to_add)

        if exp_status == "approved" and (minor_count < 3 and major_count == 0):
            if not comments and row.get("Status") != "Pending":  # Only if there are no comments
                row["Status"] = "Approved"
                row["Comments"] = "Clean MVR"

        # Excluded/Deleted logic if MVR exists
        if exp_status in ["excluded", "deleted"] and mvr_flag:
            violation_desc = str(row.get("Violation Description", "")).strip()
            # If too many minors or any major, add violations to comments
            if minor_count > 3 or major_count > 0:
                if violation_desc and violation_desc.lower() not in comments.lower():
                    row["Status"] = "Pending"
                    if comments:
                        row["Comments"] = f"{comments}, {violation_desc}"
                    else:
                        row["Comments"] = violation_desc
            # If clean, set status to Pending
            if minor_count < 3 and major_count == 0:
                pass
                row["Status"] = "Pending"

        # NEW RULE: If status is Probationary Driver or Probation, set to Pending
        # status_val = str(row.get("Status", "")).strip().lower()
        if exp_status in ["probationary driver", "probation"]:
            row["Status"] = "Pending"

        # removed deleted drivers
        filtered_data = []
        for row in mvr_data:
            exp_status = str(row.get("exp_drivers_list_status", "")).strip().lower()
            mvr_flag = bool(row.get("mvr_list_flag"))
            drv_flag = bool(row.get("driver_list_flag"))
            # If exp_status is "deleted", no MVR present, and not drv_flag, set Status as deleted
            if exp_status == "deleted" and not mvr_flag and not drv_flag:
                row['Status'] = "Deleted"
            filtered_data.append(row)

    return filtered_data

def fill_missing_dob(df, driver_df, exp_drivers_df):
    """
    Fill missing Driver DOB values by matching license numbers with driver_df and exp_drivers_df
    """
    # Create a combined reference dataframe from driver_df and exp_drivers_df
    reference_dob_data = []
    
    if driver_df is not None and not driver_df.empty:
        # Extract DOB data from driver_df
        if 'License Number' in driver_df.columns and 'Driver Date of Birth' in driver_df.columns:
            driver_ref = driver_df[['License Number', 'Driver Date of Birth']].copy()
            driver_ref = driver_ref.dropna(subset=['License Number'])
            driver_ref['Driver Date of Birth'] = pd.to_datetime(driver_ref['Driver Date of Birth'], errors='coerce')
            reference_dob_data.append(driver_ref)
    
    if exp_drivers_df is not None and not exp_drivers_df.empty:
        # Extract DOB data from exp_drivers_df
        if 'License Number' in exp_drivers_df.columns and 'Driver Date of Birth' in exp_drivers_df.columns:
            exp_ref = exp_drivers_df[['License Number', 'Driver Date of Birth']].copy()
            exp_ref = exp_ref.dropna(subset=['License Number'])
            exp_ref['Driver Date of Birth'] = pd.to_datetime(exp_ref['Driver Date of Birth'], errors='coerce')
            reference_dob_data.append(exp_ref)
    
    if not reference_dob_data:
        return df
    
    # Combine all reference data
    combined_ref = pd.concat(reference_dob_data, ignore_index=True)
    
    # Remove duplicates, keeping the first occurrence (prioritizes driver_df over exp_drivers_df)
    combined_ref = combined_ref.drop_duplicates(subset=['License Number'], keep='first')
    
    # Remove rows where DOB is null/empty
    combined_ref = combined_ref.dropna(subset=['Driver Date of Birth'])
    
    # Create a mapping dictionary for quick lookup
    license_to_dob = dict(zip(combined_ref['License Number'], combined_ref['Driver Date of Birth']))
    
    # Fill missing DOB values in df
    mask_missing_dob = df["Driver DOB"].isna() | (df["Driver DOB"] == "") | df["Driver DOB"].isnull()
    
    for idx in df[mask_missing_dob].index:
        license_num = df.loc[idx, 'license_number']
        if pd.notna(license_num) and license_num in license_to_dob:
            df.loc[idx, "Driver DOB"] = license_to_dob[license_num]
            df.loc[idx, "driver_date_of_birth"] = license_to_dob[license_num]
    
    return df

def generate_mvr_data_sheet_for_drivers(df, driver_df, vehicle_df, exp_drivers_df):
    df = df.copy()
    df = df[df['section_name'] == 'section_mvr']

    # Convert columns to datetime
    df['violation_date'] = pd.to_datetime(df['violation_date'], errors='coerce')
    df['violation_category_str'] = df['violation_category']
    df["Driver DOB"] = pd.to_datetime(df["driver_date_of_birth"], errors="coerce")
    df["license_expiration_date"] = pd.to_datetime(df["license_expiration_date"], errors="coerce")
    df["driver_hiring_date"] = pd.to_datetime(df["driver_hiring_date"], errors="coerce")
    df["mvr_generation_date"] = pd.to_datetime(df["mvr_generation_date"], errors="coerce")
    df["restrictions"] = df["restrictions"].fillna("")
    df['violation_category'] = df['violation_category'].fillna('')
    df['violation_description'] = df['violation_description'].fillna('')
    df['license_status'] = df['license_status'].fillna('')
    df['driver_first_name']= df['driver_first_name'].fillna('')
    df['driver_last_name']= df['driver_last_name'].fillna('')

    current_date = pd.Timestamp.now()
    one_year_ago = current_date - pd.Timedelta(days=365)
    three_years_ago = current_date - pd.DateOffset(years=3)
    five_years_ago = current_date - pd.DateOffset(years=5)
    df["age"] = df["Driver DOB"].apply(
        lambda dob: calculate_age(str(dob)) if pd.notnull(dob) else None
    )

    # Helper functions for aggregation
    def minor_count_last_three_years(x):
        mask = (x['violation_category'].str.lower() == 'minor') & (x['violation_date'] >= three_years_ago)
        return mask.sum()

    def minor_desc_last_three_years(x):
        mask = (x['violation_category'].str.lower() == 'minor') & (x['violation_date'] >= three_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def major_count_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= five_years_ago)
        return mask.sum()

    def major_desc_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= five_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def major_count_last_one_year(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= one_year_ago)
        return mask.sum()

    def major_desc_last_one_year(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= one_year_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def accident_count_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'accident') & (x['violation_date'] >= five_years_ago)
        return mask.sum()

    def accident_desc_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'accident') & (x['violation_date'] >= five_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def custom_major_minor_group_3_5(x):
        return custom_major_minor_3_5(
            x['violation_category_str'],
            x['violation_date'],
            three_years_ago,
            five_years_ago
        )

    # Group and aggregate
    df = fill_missing_dob(df, driver_df, exp_drivers_df)
    group_cols = ["driver_full_name", "license_number", "Driver DOB"]
    grouped = df.groupby(group_cols, sort=False)

    grouped_df = grouped.apply(lambda x: pd.Series({
        # "Driver DOB": x["Driver DOB"].iloc[0],
        "driver_first_name": x["driver_first_name"].iloc[0],
        "driver_last_name": x["driver_last_name"].iloc[0],
        "driver_hiring_date": x["driver_hiring_date"].iloc[0],
        "age": x["driver_date_of_birth"].apply(
            lambda dob: calculate_age(str(dob)) if pd.notnull(dob) else None).iloc[0],
        "mvr_generation_date": x["mvr_generation_date"].iloc[0],
        "license_state": x["license_state"].iloc[0],
        "license_class": x["license_class"].iloc[0],
        "license_status": x["license_status"].iloc[0],
        "license_expiration_date": x["license_expiration_date"].iloc[0],
        "medical_expiration_date": x["medical_expiration_date"].iloc[0],
        "medical_status": x["medical_status"].iloc[0],
        "restrictions": custom_concat(x["restrictions"]),
        "violation_category_str": custom_major_minor_group_3_5(x),
        "Minor Count": minor_count_last_three_years(x),
        "Minor Count 3 Year": minor_count_last_three_years(x),  # for post logic we will drop later
        "Minor Violation Description": minor_desc_last_three_years(x),
        "Minor Violation Description 3 Year": minor_desc_last_three_years(x),  # for post logic we will drop later
        "Major Count": major_count_last_five_years(x),
        "Major Violation Description": major_desc_last_five_years(x),
        "Major Count 1 Year": major_count_last_one_year(x),  # for post logic we will drop later
        "Major Violation Description 1 Year": major_desc_last_one_year(x),  # for post logic we will drop later
        "Major Violation Description 5 Year": major_desc_last_five_years(x),
        "Accident Count": accident_count_last_five_years(x),
        "Accident Violation Description": accident_desc_last_five_years(x),
    })).reset_index()

    # Combine all descriptions into a single column, skipping empty values
    desc_cols = [
        "Minor Violation Description",
        "Major Violation Description",
        "Accident Violation Description"
    ]
    grouped_df["violation_description"] = grouped_df[desc_cols].apply(
        lambda row: ", ".join([str(val) for val in row if val and str(val).strip() != ""]), axis=1
    )
    grouped_df = grouped_df.drop(columns=desc_cols)

    # Rename columns
    grouped_df.columns = [
        "Driver Full Name",
        "License Number",
        "Driver Date of Birth",
        # "Driver Date of Birth",
        "Driver First Name",
        "Driver Last Name",
        "Hire Date",
        "Age",
        "Date MVR Ordered",
        "License State",
        "License Type",
        "License Status",
        "License Expiration Date",
        "Medical Expiration Date",
        "Medical Status",
        "Restrictions",
        "Violation Category",
        "Minor Count",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count",
        "Major Count 1 Year",
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year",
        "Accident Count",
        "Violation Description"
    ]

    # Convert relevant columns to integers (handle NaN values)
    grouped_df["Accident Count"] = (
        pd.to_numeric(grouped_df["Accident Count"], errors="coerce")
        .apply(lambda x: int(x) if pd.notna(x) else x)
    )

    grouped_df["Minor Count"] = (
        pd.to_numeric(grouped_df["Minor Count"], errors="coerce").apply(lambda x: int(x) if pd.notna(x) else x)
    )
    grouped_df["Major Count"] = (
        pd.to_numeric(grouped_df["Major Count"], errors="coerce").apply(lambda x: int(x) if pd.notna(x) else x)
    )

    grouped_df["Total Incidents"] = (
        grouped_df[["Accident Count", "Minor Count", "Major Count"]]
        .sum(axis=1)
        .apply(lambda x: np.nan if x == 0 else int(x) if pd.notna(x) else x)
    )
    grouped_df[["Minor Count", "Major Count", "Accident Count"]] = grouped_df[
        ["Minor Count", "Major Count", "Accident Count"]
    ].replace(0, "")

    # Format 'Driver Date of Birth' and 'License Expiration Date' to 'mm/dd/yyyy'
    grouped_df["Driver Date of Birth"] = grouped_df["Driver Date of Birth"].apply(
        lambda dob: dob.strftime("%m/%d/%Y") if pd.notnull(dob) else ""
    )
    grouped_df["License Expiration Date"] = grouped_df["License Expiration Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )
    grouped_df["Hire Date"] = grouped_df["Hire Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )

    # Add columns that do not have mappings and set them to None
    grouped_df["Years of Experience"] = None
    # grouped_df["Hire Date"] = None # we will pick from drivers list
    grouped_df["Years of Tenure"] = None
    grouped_df["MVR Received"] = "FALSE"
    # grouped_df["undesirable"] = None
    grouped_df["Excluded"] = None
    # grouped_df["prohibited"] = None
    # grouped_df["mvr_score"] = None
    grouped_df["Status"] = None
    # grouped_df["Violation Description"] = None
    grouped_df["Number of Points"] = 0
    grouped_df["FullTime"] = 1
    grouped_df["Comments"] = None
    grouped_df['driver_list_flag'] = None
    grouped_df['exp_drivers_list_flag'] = None
    grouped_df['mvr_list_flag'] = None
    grouped_df['exp_drivers_list_status'] = None
    grouped_df[
        'exp_drivers_list_comments'] = None  # need to maintain extra column to keep track of exp driver list comments
    grouped_df['AddDeleteDate'] = None  # column data taken from exp
    # Reorder columns to match the required order (using label keys)
    column_order = [
        "Date MVR Ordered",
        "Driver Full Name",
        "Driver First Name",
        "Driver Last Name",
        "Driver Date of Birth",
        "Age",
        "Hire Date",
        "Years of Tenure",
        "Years of Experience",
        "MVR Received",
        "License Number",
        "License Type",
        "License State",
        "License Expiration Date",
        "License Status",
        "Status",
        "Violation Description",
        "Violation Category",
        "Accident Count",
        "Minor Count",
        "Major Count",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count 1 Year",  # only used for post logics will be dropped later
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year",  # only used for post logics will be dropped later
        "Total Incidents",
        "Excluded",
        "Number of Points",
        "FullTime",
        "Medical Expiration Date",
        "Medical Status",
        "Restrictions",
        "Comments",
        "AddDeleteDate",
        "driver_list_flag",  # this column is just for supporting post logic
        "exp_drivers_list_flag",  # this column is just for supporting post logic
        "mvr_list_flag",  # this column is just for supporting post logic
        "exp_drivers_list_status",  # this column is just for supporting post logic, will be filled from exp_drivers_df
        "exp_drivers_list_comments"  # this column is just for supporting post logic, will be filled from exp_drivers_df
    ]

    # Reorder columns
    grouped_df = grouped_df[column_order]

    if "Vehicle Body Type" in vehicle_df.columns and not vehicle_df["Vehicle Body Type"].isnull().all():
        has_heavy_vehicle = vehicle_df["Vehicle Body Type"].str.contains(r"heavy|extra heavy", case=False,
                                                                         na=False).any()
    else:
        has_heavy_vehicle = False

    grouped_df[["Status", "Comments"]] = grouped_df.apply(lambda row: status_and_comments(row, has_heavy_vehicle),
                                                          axis=1)

    grouped_df["Date MVR Ordered"] = grouped_df["Date MVR Ordered"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )
    exp_drivers_df = process_exp_driver_df(exp_drivers_df)

    # we will merge the driver df(Acord) and drivers empiring df to create the super set of drivers
    if driver_df is not None and exp_drivers_df is not None:
        # Ensure first and last name columns exist
        if not (driver_df.replace('', pd.NA).isna().all().all()):
            if 'Driver Full Name' in driver_df.columns:
                driver_df[["Driver First Name", "Driver Last Name"]] = driver_df["Driver Full Name"].apply(
                    split_driver_name)
            if 'Driver Full Name' in exp_drivers_df.columns:
                exp_drivers_df[["Driver First Name", "Driver Last Name"]] = exp_drivers_df["Driver Full Name"].apply(
                    split_driver_name)

            # we need to pick few columns from mvr output so need to rename to allow filtering on column names
            driver_df["MVR Received"] = exp_drivers_df["MVR Received"] = "FALSE"
            driver_df["Number of Points"] = exp_drivers_df["Number of Points"] = 0
            driver_df["FullTime"] = exp_drivers_df["FullTime"] = 1
            # driver_df["Comments"] = exp_drivers_df["Comments"] = "MVR Needed"
            driver_df["Status"] = exp_drivers_df["Status"] = "Pending"
            driver_df["Years of Tenure"] = driver_df["Hire Date"].apply(calculate_age)

            driver_df = driver_df.reindex(columns=column_order, fill_value="")
            exp_drivers_df = exp_drivers_df.reindex(columns=column_order, fill_value="")

            merged_drivers_df = superset_drivers(driver_df, exp_drivers_df)
            # for superset of drivers and acord, we calculate age.
            # incase mvr doesnt match then age should not be empty
            merged_drivers_df["Age"] = merged_drivers_df["Driver Date of Birth"].apply(calculate_age)

            ''' now we find similar drivers from combined driver list and mvr data '''
            # driver_df["Years of Tenure"] = driver_df["Hire Date"].apply(calculate_age)
            matched_grouped_indices = set()
            for i, driver_row in merged_drivers_df.iterrows():
                # Only consider unmatched grouped_df rows
                unmatched_grouped_df = grouped_df.loc[~grouped_df.index.isin(matched_grouped_indices)]
                # Step 1: Try to match by license number
                license_matches = unmatched_grouped_df.apply(
                    lambda grouped_row: match_by_license(driver_row, grouped_row), axis=1
                )
                matching_row = unmatched_grouped_df[license_matches]
                # Step 2: If no match, fallback to name+dob
                if matching_row.empty:
                    fallback_matches = unmatched_grouped_df.apply(
                        lambda grouped_row: match_by_name_dob(driver_row, grouped_row), axis=1
                    )
                    matching_row = unmatched_grouped_df[fallback_matches]
                if not matching_row.empty:
                    # Mark this grouped_df row as matched
                    matched_grouped_indices.add(matching_row.index[0])
                    for col in merged_drivers_df.columns:
                        # if col in ["exp_drivers_list_status"]:
                        #     pass
                        if col in ["AddDeleteDate","exp_drivers_list_flag","driver_list_flag",
                                   "exp_drivers_list_status","exp_drivers_list_comments"]:  # need to keep current value as in merged df
                            continue
                        if col == "Comments":
                            if merged_drivers_df.loc[i, 'exp_drivers_list_comments'] in ["probationary driver",
                                                                                         "probation", "excluded"]:
                                merged_drivers_df.loc[i, col] = concat_without_clean_mvr(merged_drivers_df.loc[i, col],
                                                                                         matching_row.iloc[0][col])
                            else:
                                merged_drivers_df.loc[i, col] = matching_row.iloc[0][col]

                        elif col == "mvr_list_flag":
                            merged_drivers_df.loc[i, col] = "TRUE"
                        # elif col in ["exp_drivers_list_flag","driver_list_flag"]:
                        #     pass

                        elif col == "MVR Received":
                            merged_drivers_df.loc[i, col] = "TRUE"
                        elif col == "Years of Tenure":
                            hire_date_str = matching_row.iloc[0]["Hire Date"]
                            if hire_date_str:
                                try:
                                    hire_date = parser.parse(hire_date_str).date()
                                    years_of_tenure = datetime.now().year - hire_date.year
                                    merged_drivers_df.loc[i, col] = years_of_tenure
                                except Exception:
                                    merged_drivers_df.loc[i, col] = ""
                            else:
                                merged_drivers_df.loc[i, col] = ""
                        else:
                            merged_drivers_df.loc[i, col] = matching_row.iloc[0][col]
                else:
                    # when drivers dont match, put that driver's data as empty
                    # merged_drivers_df.loc[i, "Driver Date of Birth"] = ""
                    # merged_drivers_df.loc[i, "Years of Experience"] = ""
                    # merged_drivers_df.loc[i, "License Number"] = ""
                    # merged_drivers_df.loc[i, "License State"] = ""
                    # merged_drivers_df.loc[i, "Hire Date"] = ""
                    pass

            mvr_data = merged_drivers_df.to_dict(orient="records")
            # Add unmatched grouped_df rows to mvr_data
            unmatched_grouped_df = grouped_df.loc[~grouped_df.index.isin(matched_grouped_indices)]
            if not unmatched_grouped_df.empty:
                mvr_data.extend(unmatched_grouped_df.to_dict(orient="records"))

            ''' adding post logic to final mvr '''
            mvr_data = postprocess_mvr_data(mvr_data)

            # just deleting the helper columns we added
            mvr_data = filter_and_cleanup_mvr_data(mvr_data)
        else:
            mvr_data = grouped_df.to_dict(orient="records")
    return mvr_data


def generate_report(output_dict, driver_df=None, vehicle_df=None, exp_drivers_df=None, workbook=None,
                    xlsx_sheet_tabs=[]):
    if not workbook:
        workbook = Workbook()

    result_df = pd.DataFrame(output_dict)

    mvr_output = generate_mvr_data_sheet_for_drivers(result_df, driver_df, vehicle_df, exp_drivers_df)
    
    if mvr_output:
        mvr_output_df = pd.DataFrame(mvr_output)
        try:
            #mvr_output_df_processed = extract_and_overwrite_names(mvr_output_df)
            mvr_output_df_processed = mvr_output_df
        except Exception as e:
            # print(f"Error during LLM name extraction: {e}. Proceeding with original data.")
            mvr_output_df_processed = mvr_output_df # Fallback to original data
        
        mvr_output_final = mvr_output_df_processed.to_dict(orient="records")

        write_excel_sheet_mvr_owned(
            mvr_output_final, workbook, "Riscom MVR", "Riscom MVR", False, start_row=2, skip_blue_header=True
        )

    # reorder sheets
    desired_order = ["Riscom MVR", "drivers", "vehicles", "MVR raw"]
    # Only include sheets that actually exist
    ordered_sheets = [name for name in desired_order if name in workbook.sheetnames]
    # Add any other sheets that might exist but are not in the desired order
    ordered_sheets += [name for name in workbook.sheetnames if name not in ordered_sheets]
    workbook._sheets = [workbook[name] for name in ordered_sheets]
    
   
    worksheet = workbook['MVR raw']
    workbook.active = worksheet
    report_buffer = BytesIO()
    workbook.save(report_buffer)
    report_buffer.seek(0)

    return report_buffer
def download_report_mvr_renewal_riscom_test2(uploaded_zip_file):
    # Create a BytesIO object from the uploaded zip file
    zip_buffer = io.BytesIO(uploaded_zip_file.read())

    mvr_file_buffer = None
    drivers_exp_file_buffer = None
    original_wb = None

    # Open the zip archive
    with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
        # Loop through all files in the zip
        for file_name in zip_file.namelist():
            
            # Skip system files like __MACOSX
            if file_name.startswith('__MACOSX/') or file_name.startswith('._'):
                continue

            # Get the base name (e.g., "report_mvr.xlsx") from the full path (e.g., "folder/report_mvr.xlsx")
            base_name = os.path.basename(file_name.lower())

            if base_name.startswith('report_mvr'):
                # print(f"Found MVR file: {file_name}")
                # Read the file content ONCE
                with zip_file.open(file_name) as file:
                    file_content = file.read()
                
                # Use the content to create two independent buffers
                # One for openpyxl (load_workbook) and one for pandas (read_excel)
                original_wb = load_workbook(io.BytesIO(file_content))
                mvr_file_buffer = io.BytesIO(file_content)

            elif base_name.startswith('drivers_'):
                # print(f"Found Drivers Exp file: {file_name}")
                # Read the file content ONCE
                with zip_file.open(file_name) as file:
                    file_content = file.read()
                
                # Create the buffer for pandas
                drivers_exp_file_buffer = io.BytesIO(file_content)


    # Now, check if the buffers were successfully created
    if mvr_file_buffer is None:
        raise Exception("MVR file ('report_mvr...') not found in the zip archive! Check if the file name is correct.")
    if drivers_exp_file_buffer is None:
        raise Exception("Drivers Expiring file ('drivers_...') not found in the zip archive! Check if the file name is correct.")
        
    df = pd.read_excel(mvr_file_buffer, sheet_name="MVR raw", skiprows=1, dtype={'license_number': str})
    mvr_file_buffer.seek(0) 
    driver_df = pd.read_excel(mvr_file_buffer, sheet_name="drivers", skiprows=1, dtype={'License Number': str})
    mvr_file_buffer.seek(0) 
    vehicle_df = pd.read_excel(mvr_file_buffer, sheet_name="vehicles", skiprows=1)
    
    exp_drivers_df = read_exp_drivers_excel(drivers_exp_file_buffer, sheet_name=0)

    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)


    sheets_to_keep = ["MVR raw", "drivers", "vehicles"]
    sheets_to_remove = [sheet for sheet in original_wb.sheetnames if sheet not in sheets_to_keep]
    
    for sheet_name in sheets_to_remove:
        print(f"Removing unwanted sheet: {sheet_name}")
        if sheet_name in original_wb:
            original_wb.remove(original_wb[sheet_name])
    output_dict = df.to_dict(orient="records")
    
    buffer = generate_report(output_dict, driver_df, vehicle_df, exp_drivers_df, original_wb)
    return [buffer]

if __name__ == "__main__":
    file_path = '/home/dell/Desktop/pibit/insight-board/Archive (26).zip' 
    
    try:
        with open(file_path, "rb") as uploaded_file:
            print(f"Processing file: {file_path}")

            result = download_report_mvr_renewal_riscom_test2(uploaded_file)
        
        output_filename = "RISCOM_UW_OUTPUT_LLM.xlsx"
        with open(output_filename, "wb") as f:
            f.write(result[0].getvalue())
        print(f"Successfully generated report: {output_filename}")
        
    except FileNotFoundError:
        print(f"Error: Test file not found at {file_path}")
    except Exception as e:
        print(f"An error occurred during processing: {e}")
        import traceback
        traceback.print_exc()