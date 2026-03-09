import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors

class ExcelToPdfConverter:
    """Excel to PDF converter using manual canvas drawing for absolute positioning control."""
    
    def __init__(self, input_file, output_file, dpi=300):
        self.input_file = input_file
        self.output_file = output_file
        self.dpi = dpi

    def get_real_data_range(self, ws):
        """Detect the actual used data range."""
        max_row = 0
        max_col = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)
        return max_row, max_col

    def convert(self):
        try:
            wb = openpyxl.load_workbook(self.input_file, data_only=True)
        except Exception as e:
            print(f"Error: Failed to load workbook. {e}")
            return False

        c = canvas.Canvas(self.output_file, pagesize=A4)
        page_width, page_height = A4
        margin = 36 # 0.5 inch

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row, max_col = self.get_real_data_range(ws)
            if max_row == 0 or max_col == 0:
                continue

            # --- Step 1: Calculate Scale ---
            original_col_widths = []
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                w = ws.column_dimensions[col_letter].width or 8.43
                original_col_widths.append(w * 7.0)
            
            total_content_width = sum(original_col_widths)
            available_width = page_width - (margin * 2)
            
            scale = available_width / total_content_width if total_content_width > available_width else 1.0
            if scale < 0.1: scale = 0.1 

            # --- Step 2: Render Content ---
            current_y = page_height - margin
            
            # Header
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, current_y - 12, f"Sheet: {sheet_name}")
            current_y -= 30

            for r_idx in range(1, max_row + 1):
                row_height = (ws.row_dimensions[r_idx].height or 15) * scale
                
                # Check for page break
                if current_y - row_height < margin:
                    c.showPage()
                    current_y = page_height - margin
                
                current_x = margin
                
                for c_idx, orig_width in enumerate(original_col_widths, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    col_width = orig_width * scale
                    
                    # Manual Cell Positioning with Light Padding
                    # Clean the cell area first if needed (usually canvas is fresh)
                    
                    # Optional: Draw cell background if any
                    if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            try:
                                c.setFillColor(colors.HexColor(f"#{rgb[-6:]}"))
                                c.rect(current_x, current_y - row_height, col_width, row_height, stroke=0, fill=1)
                            except: pass

                    # Draw Cell Border
                    c.setStrokeColor(colors.lightgrey)
                    c.setLineWidth(0.5 * scale)
                    c.rect(current_x, current_y - row_height, col_width, row_height, stroke=1, fill=0)
                    
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        if text:
                            font_size = max(8 * scale, 6)
                            font_name = "Helvetica"
                            if cell.font and cell.font.bold: font_name = "Helvetica-Bold"
                            
                            c.setFont(font_name, font_size)
                            c.setFillColor(colors.black)
                            
                            # Font color
                            if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
                                rgb = cell.font.color.rgb
                                if isinstance(rgb, str) and len(rgb) >= 6:
                                    try:
                                        c.setFillColor(colors.HexColor(f"#{rgb[-6:]}"))
                                    except: pass

                            padding = 2 * scale
                            text_x = current_x + padding
                            # Baseline adjustment
                            text_y = current_y - row_height + padding + (font_size * 0.15)
                            
                            # Overlap prevention: Clip text width
                            avail_txt_w = col_width - (padding * 2)
                            if c.stringWidth(text, font_name, font_size) > avail_txt_w:
                                txt_draw = text
                                while len(txt_draw) > 0 and c.stringWidth(txt_draw + "...", font_name, font_size) > avail_txt_w:
                                    txt_draw = txt_draw[:-1]
                                txt_draw += "..."
                            else:
                                txt_draw = text

                            c.drawString(text_x, text_y, txt_draw)

                    current_x += col_width
                
                current_y -= row_height

            c.showPage() 

        wb.close()
        try:
            c.save()
            print(f"Successfully created {self.output_file} with Manual Position Tracking.")
            return True
        except Exception as e:
            print(f"Error building PDF: {e}")
            return False

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_to_pdf_converter.py <input_xlsx> <output_pdf>")
    else:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
        if not os.path.exists(input_file):
            print(f"Error: File {input_file} not found.")
            sys.exit(1)
        converter = ExcelToPdfConverter(input_file, output_file)
        converter.convert()
