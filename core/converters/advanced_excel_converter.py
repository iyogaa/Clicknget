import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from typing import List, Tuple, Optional, Dict, Any
import io
import os

from core.exceptions.exceptions import ConversionError
from core.utils.logger import logger

class ExcelToPdfConverter:
    """Advanced logic for converting Excel files to PDF using manual canvas positioning to prevent overlapping."""
    
    def __init__(self, dpi: int = 300):
        self.dpi = dpi

    def _get_clean_data_range(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[int, int]:
        max_row = 0
        max_col = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)
        return max_row, max_col

    def convert(self, input_source: Any, output_path: str, options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Convert Excel to PDF using Manual Position Tracking and Proportional Scaling.
        """
        options = options or {}
        p_size = A4
        margin = 36 # 0.5 inch
        
        try:
            if isinstance(input_source, bytes):
                wb = openpyxl.load_workbook(io.BytesIO(input_source), data_only=True)
            else:
                wb = openpyxl.load_workbook(input_source, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise ConversionError(f"Failed to load workbook: {e}")

        c = canvas.Canvas(output_path, pagesize=p_size)
        pw, ph = p_size
        avail_w = pw - (margin * 2)

        sheets_processed = []
        warnings = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row, max_col = self._get_clean_data_range(ws)
            if max_row == 0 or max_col == 0:
                continue

            sheets_processed.append(sheet_name)
            
            # --- Dimensions & Scaling ---
            original_col_widths = []
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                w = ws.column_dimensions[col_letter].width or 8.43
                original_col_widths.append(w * 7.0)
            
            total_content_width = sum(original_col_widths)
            scale = avail_w / total_content_width if total_content_width > avail_w else 1.0
            if scale < 0.1:
                scale = 0.1
                warnings.append(f"Sheet '{sheet_name}' hit 10% scale floor.")

            # --- Drawing ---
            current_y = ph - margin
            
            # Header
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, current_y - 12, f"Sheet: {sheet_name}")
            current_y -= 30

            for r_idx in range(1, max_row + 1):
                row_height = (ws.row_dimensions[r_idx].height or 15) * scale
                
                # Page break check
                if current_y - row_height < margin:
                    c.showPage()
                    current_y = ph - margin
                
                current_x = margin
                
                for c_idx, orig_w in enumerate(original_col_widths, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    col_width = orig_w * scale
                    
                    # Optional cell fill
                    if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            try:
                                c.setFillColor(colors.HexColor(f"#{rgb[-6:]}"))
                                c.rect(current_x, current_y - row_height, col_width, row_height, stroke=0, fill=1)
                            except: pass

                    # Draw Cell Border
                    c.setStrokeColor(colors.lightgrey)
                    c.setLineWidth(0.5 * scale)
                    c.rect(current_x, current_y - row_height, col_width, row_height, stroke=1, fill=0)
                    
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        if text:
                            font_size = max(8 * scale, 6)
                            font_name = "Helvetica"
                            if cell.font and cell.font.bold: font_name = "Helvetica-Bold"
                            
                            c.setFont(font_name, font_size)
                            c.setFillColor(colors.black)
                            
                            if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
                                rgb = cell.font.color.rgb
                                if isinstance(rgb, str) and len(rgb) >= 6:
                                    try:
                                        c.setFillColor(colors.HexColor(f"#{rgb[-6:]}"))
                                    except: pass

                            padding = 2 * scale
                            text_x = current_x + padding
                            text_y = current_y - row_height + padding + (font_size * 0.15)
                            
                            # Clip text to cell width
                            avail_txt_w = col_width - (padding * 2)
                            if c.stringWidth(text, font_name, font_size) > avail_txt_w:
                                txt_draw = text
                                while len(txt_draw) > 0 and c.stringWidth(txt_draw + "...", font_name, font_size) > avail_txt_w:
                                    txt_draw = txt_draw[:-1]
                                txt_draw += "..."
                            else:
                                txt_draw = text

                            c.drawString(text_x, text_y, txt_draw)

                    current_x += col_width
                
                current_y -= row_height

            c.showPage()

        wb.close()
        try:
            c.save()
            return {
                "success": True,
                "sheets_processed": sheets_processed,
                "warnings": warnings,
                "output_path": output_path
            }
        except Exception as e:
            logger.error(f"Canvas Save Error: {e}")
            raise ConversionError(f"Canvas Save Error: {e}")

def convert_excel_to_pdf(input_path: str, output_path: str, **kwargs) -> Dict[str, Any]:
    """Production entry point for Excel to PDF converter."""
    converter = ExcelToPdfConverter()
    return converter.convert(input_path, output_path, options=kwargs)
