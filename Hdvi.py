import datetime
import re
import yaml
import openpyxl
import pandas as pd
import numpy as np

from fuzzywuzzy import fuzz
from dateutil import parser
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

blue_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
blue_header_fill = PatternFill(
    start_color="0033CCCC", end_color="0033CCCC", fill_type="solid"
)
def read_state_mapping(file_path="state_mapping.yaml"):
    with open(file_path, 'r') as file:
        state_mapping = yaml.safe_load(file)
    return state_mapping


def strip_non_numeric_and_leading_zeros(input_string):
    numeric_part = re.sub(r"\D", "", str(input_string))
    return numeric_part.lstrip("0")

def format_date_cell(cell):
    try:
        if pd.isna(cell):  # Handle NaN values safely
            return pd.NaT
        if isinstance(cell, pd.Timestamp) or isinstance(cell, datetime.datetime):  # If already a Timestamp, return formatted string
            return cell.strftime('%m/%d/%Y')
        cell = str(cell).strip()  # Convert to string and strip spaces
        cell = cell.replace("-", "/")
        # Try parsing with 4-digit year first
        date = pd.to_datetime(cell, format='%m/%d/%Y', errors='coerce')
        if pd.isna(date):
            # If parsing fails, try with 2-digit year
            date = pd.to_datetime(cell, format='%m/%d/%y', errors='coerce')
            if date.year > 2025:
                date = date.replace(year=date.year - 100)
        
        return date.strftime('%m/%d/%Y') if pd.notna(date) else pd.NaT  # Standard output format
    except Exception as e:
        print(f"Error converting date: {e}")
        return cell

def filter_drivers(driver_row, grouped_row):
    name_driver = f"{driver_row['First Name']} {driver_row['Last Name']}"
    name_grouped = f"{grouped_row['First Name']} {grouped_row['Last Name']}"
    cdl_driver, cdl_grouped = driver_row["CDL Number"], grouped_row["CDL Number"]
    dob_driver, dob_grouped = driver_row["Date of Birth"], grouped_row["Date of Birth"]

    if strip_non_numeric_and_leading_zeros(
        cdl_driver
    ) == strip_non_numeric_and_leading_zeros(cdl_grouped) and not pd.isnull(cdl_driver) and cdl_driver is not None and cdl_driver != "":
        return True
    try:
        dob_driver_parsed = parser.parse(dob_driver).date()
        dob_grouped_parsed = parser.parse(dob_grouped).date()
    except (ValueError, TypeError):
        dob_driver_parsed = None
        dob_grouped_parsed = None
    if dob_driver_parsed and dob_grouped_parsed and dob_driver_parsed is not None:
        if dob_driver_parsed == dob_grouped_parsed:
            return match_driver_names(name_driver, name_grouped) >= 80
    else:
        return match_driver_names(name_driver, name_grouped) >= 80
    
    return False


def match_driver_names(name_driver, name_grouped):
    if name_driver != "":
        set_ratio = fuzz.token_set_ratio(name_driver, name_grouped)
        sort_ratio = fuzz.token_sort_ratio(name_driver, name_grouped)
        return (set_ratio + sort_ratio) / 2
    return False

def generate_mvr_data_sheet(df, driver_df):
    # Convert date columns to datetime format
    df["Driver DOB"] = pd.to_datetime(df["Driver Date of Birth"], errors="coerce")
    df["Expiration Date"] = pd.to_datetime(
        df["License Expiration Date"], errors="coerce"
    )

    # Define current date
    current_date = pd.Timestamp.now()

    # Calculate 'Age' from 'Driver DOB'
    df["Age"] = df["Driver DOB"].apply(
        lambda dob: current_date.year - dob.year if pd.notnull(dob) else None
    )

    df.fillna({col: "" for col in df.columns}, inplace=True)  # Fill None with empty strings
    
    grouped_df = (
        df.groupby(
            ["Driver Full Name", "CDL Number", "Driver Date of Birth"], sort=False
        )
        .agg(
            {
                "Driver DOB": "first",
                "Driver First Name": "first",  # First Name
                "Driver Last Name": "first",  # Last Name
                "Age": "first",  # Age (calculated)
                "License State": "first",  # License State
                "CDL Type": "first",  # CDL Type
                "Expiration Date": "first",  # Expiration Date
                "Violation Category": lambda x: (
                    (x == "Minor").sum(),  # Minor Count
                    (x == "Major").sum(),  # Major Count
                    (x == "Prohibited").sum(),  # Prohibited Count
                    (x == "Accident").sum(),  # Accident Count
                ),
            }
        )
        .reset_index()
    )

    # Rename columns
    grouped_df.columns = [
        "Driver Full Name",
        "CDL Number",
        "DOB",
        "Date of Birth",
        "First Name",
        "Last Name",
        "Age",
        "License State",
        "CDL Type",
        "Expiration Date",
        "Violation Category Counts",
    ]

    # Split 'Violation Category Counts' into separate columns
    grouped_df[["Minor Count", "Major Count", "Prohibited Count", "Accident Count"]] = (
        pd.DataFrame(
            grouped_df["Violation Category Counts"].tolist(), index=grouped_df.index
        )
    )
    grouped_df[["Minor Count", "Major Count", "Prohibited Count", "Accident Count"]] = (
        grouped_df[
            ["Minor Count", "Major Count", "Prohibited Count", "Accident Count"]
        ].replace(0, "")
    )

    # Convert relevant columns to integers (handle NaN values)
    grouped_df["Accident Count"] = pd.to_numeric(
        grouped_df["Accident Count"], errors="coerce"
    ).apply(lambda x: int(x) if pd.notna(x) else x)
    grouped_df["Minor Count"] = pd.to_numeric(
        grouped_df["Minor Count"], errors="coerce"
    ).apply(lambda x: int(x) if pd.notna(x) else x)
    grouped_df["Major Count"] = pd.to_numeric(
        grouped_df["Major Count"], errors="coerce"
    ).apply(lambda x: int(x) if pd.notna(x) else x)
    grouped_df["Prohibited Count"] = pd.to_numeric(
        grouped_df["Prohibited Count"], errors="coerce"
    ).apply(lambda x: int(x) if pd.notna(x) else x)

    # Calculate 'Total Incidents' (sum of all counts)
    grouped_df["Total Incidents"] = (
        grouped_df[["Accident Count", "Minor Count", "Major Count", "Prohibited Count"]]
        .sum(axis=1)
        .apply(lambda x: np.nan if x == 0 else int(x) if pd.notna(x) else x)
    )
    # Drop the 'Violation Category Counts' column
    grouped_df.drop(columns=["Violation Category Counts"], inplace=True)

    # Format 'Date of Birth' to 'mm/dd/yyyy'
    grouped_df["Date of Birth"] = grouped_df["Date of Birth"].apply(
    lambda dob: dob.strftime("%m/%d/%Y") if pd.notnull(dob) else ""
    )

    grouped_df["Expiration Date"] = grouped_df["Expiration Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )

    # Add columns that do not have mappings and set them to None
    grouped_df["Years of Experience"] = None
    grouped_df["Hire Date"] = None
    grouped_df["Years of Tenure"] = None
    grouped_df["MVR Received"] = [False]*grouped_df.shape[0]
    grouped_df["Undesirable"] = None
    grouped_df["Excluded"] = None
    grouped_df["Prohibited"] = [False]*grouped_df.shape[0]
    grouped_df["MVR Score"] = None

    # Reorder columns to match the required order
    column_order = [
        "First Name",
        "Last Name",
        "Date of Birth",
        "Age",
        "Years of Experience",
        "Hire Date",
        "Years of Tenure",
        "License State",
        "CDL Number",
        "CDL Type",
        "Expiration Date",
        "MVR Received",
        "Accident Count",
        "Minor Count",
        "Major Count",
        "Prohibited Count",
        "Total Incidents",
        "Undesirable",
        "Excluded",
        "Prohibited",
        "MVR Score",
    ]
    state_mapping = read_state_mapping()
    # Reorder columns
    grouped_df = grouped_df[column_order]
    for column in column_order:
        if column not in driver_df.columns:
            driver_df[column] = None
    driver_df = driver_df[column_order]
    if "First Name" not in driver_df.columns and "Last Name" not in driver_df.columns and "Name" in driver_df.columns:
        driver_df["First Name"] = driver_df.apply(
            lambda row: (
                row["First Name"]
                if pd.notnull(row["First Name"])
                else row["Name"].split(" ", 1)[0]
            ),
            axis=1,
        )
        driver_df["Last Name"] = driver_df.apply(
            lambda row: (
                row["Last Name"]
                if pd.notnull(row["Last Name"])
                else row["Name"].split(" ", 1)[1] if " " in row["Name"] else ""
            ),
            axis=1,
        )
    elif "Name" in driver_df.columns:
        driver_df[["First Name", "Last Name"]] = driver_df["Name"].str.split(
            " ", n=1, expand=True
        )
    driver_df["Date of Birth"] =  driver_df["Date of Birth"].apply(lambda cell: format_date_cell(cell))
    driver_df["Expiration Date"] = driver_df["Expiration Date"].apply(lambda cell: format_date_cell(cell))
    driver_df["Hire Date"] = driver_df["Hire Date"].apply(lambda cell: format_date_cell(cell))
    
    
    try:
        driver_df["License State"] = driver_df["License State"].str.lower().map(state_mapping).fillna(driver_df["License State"])
    except Exception as e:
        print(f"Error mapping license states: {e}")
        driver_df["License State"] = None
    for i, driver_row in driver_df.iterrows():
        try:
            matching_rows = grouped_df.apply(
                lambda grouped_row: filter_drivers(driver_row, grouped_row), axis=1
            )
            matching_row = grouped_df[matching_rows]
            if not matching_row.empty:
                driver_df.loc[i, "MVR Received"] = True
                driver_df.loc[i, "Prohibited"] = True
                for col in driver_df.columns:
                    if driver_df.loc[i, col] == "" or driver_df.loc[i, col] is None or pd.isna(driver_df.loc[i, col]):  # Check for empty string, None, or NaN in column
                        driver_df.loc[i, col] = matching_row.iloc[0][col]
            else:
                driver_df.loc[i, "MVR Received"] = False
                driver_df.loc[i, "Prohibited"] = False
        except Exception as e:
            print(f"Error processing driver row {i}: {e}")
            continue
    
    driver_df["Date of Birth"] = driver_df["Date of Birth"].apply(
    lambda dob: dob.strftime("%m/%d/%Y") if pd.notnull(dob) and isinstance(dob, pd.Timestamp) else dob
    )
    #Hire date issue (Century logic)
    # 🔹 Fix Hire Date parsing properly first
    driver_df["Hire Date"] = pd.to_datetime(
        driver_df["Hire Date"],
        errors="coerce"
    )

    def fix_century(date):
        if pd.isna(date):
            return pd.NaT
        if date.year < 1950:   # handles 1926 → 2026
            return date.replace(year=date.year + 100)
        return date

    driver_df["Hire Date"] = driver_df["Hire Date"].apply(fix_century)

    # 🔹 Final formatting
    driver_df["Hire Date"] = driver_df["Hire Date"].apply(
        lambda x: x.strftime("%m/%d/%Y") if pd.notnull(x) else ""
    )
    driver_df["Expiration Date"] = driver_df["Expiration Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd)  and isinstance(expd, pd.Timestamp) else expd
        )
    mvr_data = driver_df.to_dict(orient="records")
    return mvr_data


def apply_formatting_mvr_owned(
    worksheet, min_col, max_col, start_row=None, end_row=None
):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(
        min_col=min_col, max_col=max_col, min_row=start_row - 1, max_row=end_row
    ):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 2:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row >= start_row and cell.row <= end_row and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            cell.number_format = "0"
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                

def add_blue_header(sheet, header_text, last_column=5, start_row=None):
    """

    :param sheet:
    :param header_text:
    :param last_column:
    :return:
    """
    if not start_row:
        sheet.insert_rows(1)
    elif start_row and start_row == 2:
        sheet.insert_rows(1)
    else:
        sheet.insert_rows(start_row - 2)

    if start_row:
        start_row = start_row - 1
    else:
        start_row = start_row if start_row else 1

    sheet.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=last_column
    )
    header_cell = sheet.cell(row=start_row, column=1)
    header_cell.value = header_text
    header_cell.font = blue_header_font  # Font with smaller size
    header_cell.fill = blue_header_fill
    header_cell.alignment = Alignment(horizontal="center")


def write_excel_sheet_mvr_owned(
    data1, workbook, sheet_name, header_text, _first_sheet, start_row=2
):
    # global _first_sheet
    # charts = data[1]
    start_row_real = start_row
    data = data1
    if isinstance(data, dict):
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # creating tables for the graph data
        add_blue_header(worksheet, header_text, last_column=2, start_row=start_row_real)
        # if start_row != 2:
        # start_row = start_row-1
        start_row = start_row
        start_column = 1
        for row in data.items():
            for col_index, value in enumerate(row, start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        apply_formatting_mvr_owned(
            worksheet,
            min_col=1,
            max_col=2,
            start_row=start_row_real,
            end_row=start_row_real + len(data) - 1,
        )

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].auto_size = True

    elif isinstance(data, list):
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)
        start_row = start_row - 1
        start_column = 1
        last_column = 5

        header_row_flag = True
        for row in data:
            last_column = len(row)
            if header_row_flag:
                header_row_flag = False
                for col_index, value in enumerate(row.values(), start=start_column):
                    worksheet.cell(
                        row=start_row,
                        column=col_index,
                        value=list(row.keys())[col_index - 1],
                    )
                start_row += 1
            for col_index, value in enumerate(row.values(), start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        add_blue_header(worksheet, header_text, last_column, start_row=start_row_real)
        apply_formatting_mvr_owned(
            worksheet,
            min_col=1,
            max_col=len(data[0]),
            start_row=start_row_real,
            end_row=start_row_real + len(data),
        )


def generate_mvr_excel_sheets(mvr_df, client_df):
    workbook = openpyxl.Workbook()
    write_excel_sheet_mvr_owned(
        mvr_df.to_dict(orient="records"), workbook, "MVR", "MVR", True, start_row=2
    )
    modified_driver_df = generate_mvr_data_sheet(mvr_df, client_df)
    write_excel_sheet_mvr_owned(
        modified_driver_df, workbook, "hdvi output", "hdvi_output", False
    )

    return workbook    

