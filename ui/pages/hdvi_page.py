import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from unittest import mock
from core.processors.hdvi_processor import generate_mvr_excel_sheets


def remove_familial_suffixes(text):
    pattern = r'\b(Jr\.|Sr\.|I{1,3}|IV|V|VI|VII|VIII|IX|X)\b'
    cleaned_text = re.sub(pattern, '', text)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
    return cleaned_text


def get_name_component(text, position):
    text = remove_familial_suffixes(str(text))
    text = re.sub(r'[^\w\s]', '', text)

    if position == "first":
        return text.split(" ")[0]

    if position == "last":
        return text.split(" ")[-1] if " " in text else ""

    return text


def run_hdvi_mvr():
    if "hdvi_mapping_done" not in st.session_state:
        st.session_state.hdvi_mapping_done = False
    st.markdown("Upload Client Excel/CSV")
    client_file = st.file_uploader(
        "Upload Client Excel/CSV",
        type=["xlsx", "xls", "csv"],
        key="hdvi_client",
        label_visibility="collapsed"
    )

    st.markdown("Upload Output Excel")
    excel_file = st.file_uploader(
        "Upload Output Excel (containing MVR sheet)",
        type=["xlsx", "xls"],
        key="hdvi_excel",
        label_visibility="collapsed"
    )

    client_df = None
    output_file_name = "HDVI_Output.xlsx"

    if excel_file is not None:
        output_file_name = excel_file.name
        if output_file_name.startswith("report_"):
            output_file_name = output_file_name[len("report_"):]

    if client_file is not None:

        skip_rows = st.text_input(
            "Number of rows to skip",
            value="0",
            help="Skip header rows if needed"
        )

        try:
            skip_rows_int = int(skip_rows) if skip_rows else 0
        except ValueError:
            skip_rows_int = 0

        if client_file.name.endswith(".csv"):

            client_df = pd.read_csv(client_file, skiprows=skip_rows_int)
            client_df.replace("", pd.NA, inplace=True)
            client_df.dropna(how="all", inplace=True)

        elif client_file.name.lower().endswith((".xlsx", ".xls")):

            try:
                preview_wb = openpyxl.load_workbook(client_file, read_only=True)
                sheets = preview_wb.sheetnames

                sheet_name = st.selectbox(
                    "Select sheet to read from Client Excel",
                    options=sheets
                )

                client_file.seek(0)

                if sheet_name:
                    client_df = pd.read_excel(
                        client_file,
                        sheet_name=sheet_name,
                        skiprows=skip_rows_int
                    )

            except Exception as e:
                st.error(f"❌ Error reading Client Excel: {e}")

    mapping = {}
    if "hdvi_mapping_done" not in st.session_state:
        st.session_state.hdvi_mapping_done = False
    columns_available = False

    required_columns = [
        "First Name",
        "Last Name",
        "Date of Birth",
        "Age",
        "Years of Experience",
        "Hire Date",
        "Years of Tenure",
        "License State",
        "CDL Number",
        "Expiration Date"
    ]

    if client_df is not None:

        columns_available = all(col in client_df.columns for col in required_columns)

        if not columns_available:

            column_missing = list(set(required_columns) - set(client_df.columns))

            st.warning("⚠ Default column names not found. Please map alternate columns.")

            mapping["Full Name"] = st.selectbox(
                "Full Name column (use if First/Last Name not available)",
                [""] + list(client_df.columns)
            )

            for column in column_missing:
                mapping[column] = st.selectbox(
                    f"Column for {column}",
                    [""] + list(client_df.columns)
                )

            if st.button("Apply Column Mapping"):
                st.session_state.hdvi_mapping_done = True
            #Applying column mapping without clicking button 
            if columns_available or st.session_state.hdvi_mapping_done:
                if mapping["Full Name"]:

                    client_df["First Name"] = client_df[mapping["Full Name"]].apply(
                        lambda x: get_name_component(x, "first") if pd.notna(x) else ""
                    )

                    client_df["Last Name"] = client_df[mapping["Full Name"]].apply(
                        lambda x: get_name_component(x, "last") if pd.notna(x) else ""
                    )

                    mapping.pop("Full Name", None)

                for key, value in mapping.items():

                    if value:
                        client_df[key] = client_df[value]
                    else:
                        client_df[key] = None

                client_df = client_df[required_columns]

    if client_df is not None and excel_file is not None:

        if columns_available or st.session_state.hdvi_mapping_done:

            col1, col2, col3 = st.columns([2, 1.5, 2])

            with col2:

                if st.button("Generate Report", use_container_width=True):

                    try:

                        with st.spinner("Processing HDVI Report..."):

                            p = mock.patch(
                                "openpyxl.styles.fonts.Font.family.max",
                                new=100
                            )
                            p.start()

                            excel_file.seek(0)

                            with io.BytesIO(excel_file.read()) as excel_bytes:

                                workbook_mvr = openpyxl.load_workbook(
                                    excel_bytes,
                                    data_only=True
                                )

                                if "MVR" in workbook_mvr.sheetnames:

                                    mvr_sheet = workbook_mvr["MVR"]
                                    data = list(mvr_sheet.values)

                                    if data and len(data) > 1:

                                        mvr_df = pd.DataFrame(
                                            data[2:],
                                            columns=data[1]
                                        )

                                    else:
                                        mvr_df = pd.DataFrame()

                                else:

                                    st.error(
                                        "❌ The uploaded Output Excel does not contain an 'MVR' sheet."
                                    )
                                    st.stop()

                            client_df.replace("", pd.NA, inplace=True)
                            client_df.dropna(how="all", inplace=True)

                            final_wb = generate_mvr_excel_sheets(
                                mvr_df,
                                client_df
                            )

                            excel_out_bytes = io.BytesIO()
                            final_wb.save(excel_out_bytes)
                            excel_out_bytes.seek(0)

                            st.download_button(
                                label="Download",
                                data=excel_out_bytes,
                                file_name=output_file_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )

                            p.stop()

                    except Exception as e:

                        st.error(f"❌ Error generating report: {e}")

                        if st.checkbox("Show technical details"):

                            import traceback
                            st.code(traceback.format_exc())