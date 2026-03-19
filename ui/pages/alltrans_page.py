import streamlit as st
import io
import pandas as pd
from openpyxl import load_workbook
from core.processors.alltrans_processor import Alltrans


def run_mvr_all_trans():
    tab1, tab2 = st.tabs(["Upload & Process", "Help"])

    with tab1:
        st.markdown("### Upload Files")

        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### Client Excel (Lookup Data)")
            lookup_file = st.file_uploader(
                "Upload Client Excel",
                type=["xlsx", "xls", "csv"],
                key="alltrans_client",
                label_visibility="collapsed"
            )

        with col2:
            st.markdown("#### MVR File")
            main_file = st.file_uploader(
                "Upload MVR File",
                type=["xlsx", "xls", "csv"],
                key="alltrans_mvr",
                label_visibility="collapsed"
            )

        if main_file and lookup_file:
            # ---------- READ LOOKUP FILE FOR COLUMN PREVIEW ----------
            file_name = lookup_file.name.lower()

            try:
                if file_name.endswith(".csv"):
                    lookup_preview = pd.read_csv(io.BytesIO(lookup_file.getvalue()))
                else:
                    lookup_preview = pd.read_excel(io.BytesIO(lookup_file.getvalue()))

                lookup_columns = list(lookup_preview.columns)

            except Exception as e:
                st.error(f"Failed to read lookup file: {e}")
                return

            # ---------- SHEET SELECTION ----------
            chosen_sheet = None

            try:
                lookup_wb = load_workbook(
                    io.BytesIO(lookup_file.getvalue()),
                    read_only=True
                )

                sheets = lookup_wb.sheetnames

                if len(sheets) > 1:
                    chosen_sheet = st.selectbox(
                        "Lookup workbook has multiple sheets. Choose one:",
                        options=sheets,
                        key="sheet_selector"
                    )
                else:
                    chosen_sheet = sheets[0]
                    st.info(f"✅ Detected sheet: **{chosen_sheet}**")

            except Exception:
                st.info("📄 CSV or non-Excel file detected")

            st.divider()

            # ---------- COLUMN MAPPING UI ----------
            st.markdown("### Column Mapping (Optional)")

            # Read the actual lookup file to get columns
            try:
                if file_name.endswith(('.xlsx', '.xls')):
                    actual_lookup_df = pd.read_excel(
                        io.BytesIO(lookup_file.getvalue()), 
                        sheet_name=chosen_sheet if chosen_sheet else 0
                    )
                else:
                    actual_lookup_df = pd.read_csv(io.BytesIO(lookup_file.getvalue()))
                
                lookup_columns = list(actual_lookup_df.columns)
            except Exception as e:
                st.error(f"Could not read columns from lookup file: {e}")
                return

            auto_detect = st.checkbox("Auto-detect columns", value=True)

            column_map = None

            if not auto_detect:
                st.markdown("#### Manual Column Mapping")
                st.caption("Select columns from your data. Leave as '(Auto-detect)' to use automatic detection.")
                
                # Create a more comprehensive column selection
                col_options = ["(Auto-detect)"] + lookup_columns
                
                # Create a grid layout for better organization
                col1, col2, col3 = st.columns(3)
                col4, col5, col6 = st.columns(3)
                
                with col1:
                    cdl_col = st.selectbox(
                        "CDL/License Number Column",
                        options=col_options,
                        key="cdl_column",
                        help="Column containing driver license numbers"
                    )
                
                with col2:
                    hire_col = st.selectbox(
                        "Hire Date Column", 
                        options=col_options,
                        key="hire_column",
                        help="Column containing hire/employment dates"
                    )
                
                with col3:
                    name_col = st.selectbox(
                        "Name Column",
                        options=col_options,
                        key="name_column",
                        help="Column containing driver names"
                    )
                
                with col4:
                    dob_col = st.selectbox(
                        "Date of Birth Column",
                        options=col_options,
                        key="dob_column",
                        help="Column containing dates of birth"
                    )
                
                with col5:
                    state_col = st.selectbox(
                        "State Column",
                        options=col_options,
                        key="state_column",
                        help="Column containing license states (optional)"
                    )
                
                with col6:
                    # Add any other important columns you want to map
                    other_cols = [col for col in lookup_columns if col not in [cdl_col, hire_col, name_col, dob_col, state_col] and col != "(Auto-detect)"]
                    other_col_options = ["(None)"] + other_cols if other_cols else ["(None)"]
                    other_col = st.selectbox(
                        "Additional Column",
                        options=other_col_options,
                        key="other_column",
                        help="Any other important column"
                    )
                
                # Build comprehensive column map
                column_map = {}
                if cdl_col and cdl_col != "(Auto-detect)":
                    column_map["cdl"] = cdl_col
                if hire_col and hire_col != "(Auto-detect)":
                    column_map["hire_date"] = hire_col
                if name_col and name_col != "(Auto-detect)":
                    column_map["name"] = name_col
                if dob_col and dob_col != "(Auto-detect)":
                    column_map["dob"] = dob_col
                if state_col and state_col != "(Auto-detect)":
                    column_map["state"] = state_col
                if other_col and other_col != "(None)":
                    column_map["other"] = other_col
                
                # Show what will be used
                if column_map:
                    st.info(f"Selected columns: {', '.join([f'{k}={v}' for k, v in column_map.items() if v and v != '(Auto-detect)' and v != '(None)'])}")

            st.divider()

            # ---------- PROCESS BUTTON ----------
            col1, col2, col3 = st.columns([2, 1.5, 2])

            with col2:
                if st.button("Generate Report", use_container_width=True, type="primary"):
                    try:
                        # Reset file pointers
                        main_file.seek(0)
                        lookup_file.seek(0)
                        
                        main_bytes = main_file.read()
                        lookup_bytes = lookup_file.read()

                        gen = Alltrans(
                            template_path="Template.xlsx",
                            alltrans_sheet="All Trans",
                            alltrans_header_row=4,
                            mvr_sheet_name="MVR",
                            interactive_mode=False
                        )

                        with st.spinner("Processing Alltrans..."):
                            out = gen.run(
                                main_bytes,
                                lookup_bytes,
                                chosen_lookup_sheet=chosen_sheet,
                                preview_rows=8,
                                column_map=column_map
                            )

                        original_name = getattr(main_file, "name", None)

                        base = (
                            original_name.rsplit(".", 1)[0]
                            if original_name
                            else "Final_Report"
                        )

                        out_name = f"{base}.xlsx"

                        st.success("✅ Report generated successfully!")
                        
                        st.download_button(
                            "📥 Download Report",
                            data=out,
                            file_name=out_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    except Exception as e:
                        st.error(f"❌ Error: {e}")

                        if st.checkbox("Show technical details"):
                            import traceback
                            st.code(traceback.format_exc())

    # ---------- HELP TAB ----------
    with tab2:
        st.markdown(
        """
        ### 📖 How to Use Alltrans Processing

        **Step 1: Prepare Your Files**

        - **MVR File**: Motor Vehicle Record data  
        - **Client File**: Lookup file with client data

        **Step 2: Upload Files**

        1. Upload MVR file  
        2. Upload client file  
        3. Choose sheet if Excel has multiple sheets

        **Step 3: Column Mapping**

        - Leave **Auto-detect enabled** if headers match standard names
        - Disable it to manually select columns when auto-detection fails

        **Step 4: Generate Report**

        - Click **Generate Report**
        - Wait for processing to complete
        - Download the final report

        ### 💡 Tips

        - Ensure lookup file has correct column headers
        - Avoid duplicate CDL numbers in client data
        - Keep consistent formatting across files
        - Check that date formats are readable

        ### 🔄 Supported Columns

        The system can auto-detect these column types:
        - **CDL/License**: "CDL Number", "License Number", etc.
        - **Name**: "Employee Name", "Driver Name", etc.  
        - **DOB**: "Date of Birth", "DOB", etc.
        - **Hire Date**: "Hire Date", "Employment Date", etc.
        - **State**: "License State", "State", etc.
        """
        )

        st.divider()
        
        st.markdown("### 🛠️ Troubleshooting")
        
        st.markdown("""
        **Common Issues:**
        
        1. **Names not matching**: Check name formats between MVR and client files
        2. **Dates not parsing**: Ensure dates are in standard MM/DD/YYYY format
        3. **Missing drivers**: Verify CDL numbers match between files
        4. **Slow processing**: Large files may take longer to process
        
        **Need Help?**
        - Contact support with your files and error messages
        - Check that Template.xlsx exists in the correct location
        """)

if __name__ == "__main__":
    run_mvr_all_trans()
