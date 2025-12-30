import pandas as pd
import streamlit as st
import io
import os
import tempfile
from Pdf_maker import process_pdf

st.set_page_config(layout="wide")

st.markdown("""
<style>
.stApp {
    background-color: #000000;
}
[data-testid="stSidebar"] {
    background-color: #111111 !important;
}
/* Sidebar text bright white but now in normal font */
[data-testid="stSidebar"] * {
    color: #FFFFFF !important;
    font-weight: normal;
}
/* Left-aligned Heading with smaller font and no border */
.custom-heading {
    font-size: 2rem;
    color: white;
    text-align: left;
    font-weight: bold;
    margin-bottom: 1.5rem;
    margin-left: 2rem;
    background: none;
    border: none;
    padding: 0;
}
/* Remove extra empty box inside file uploader */
[data-testid="stFileUploader"] > div {
    background-color: transparent !important;
    padding: 0 !important;
    margin: 0 !important;
    border: none !important;
    min-height: 0 !important;
    min-width: 0 !important;
}
/* Label and input text white */
label, .stFileUploader, .stNumberInput label, .stSelectbox label {
    color: white !important;
}
/* White text for all content */
body, .stMarkdown, .stText, .stDataFrame, .stMetric {
    color: white !important;
}
/* Custom button styling */
.stButton>button {
    background-color: #000000;
    color: white;
    border-radius: 5px;
    padding: 0.5rem 1rem;
    font-weight: bold;
}
/* Status indicator */
.status-indicator {
    display: inline-block;
    width: 10px;
    height: 10px;
    border-radius: 50%;
    margin-right: 5px;
}
.status-operational {
    background-color: #4CAF50;
}
</style>
""", unsafe_allow_html=True)

def authenticate(username, password):
    if "credentials" in st.secrets and username in st.secrets["credentials"]:
        user_data = st.secrets["credentials"][username]
        if password == user_data["password"]:
            return user_data["role"]
    return None

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["username"] = None
    st.session_state["role"] = None

def show_login():
    with st.sidebar:
        st.title("Login")
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("Login"):
            role = authenticate(username, password)
            if role:
                st.session_state["authenticated"] = True
                st.session_state["username"] = username
                st.session_state["role"] = role
                st.rerun()
            else:
                st.error("Invalid username or password")

if not st.session_state["authenticated"]:
    show_login()
    st.stop()

def get_menu_options(role):
    base = ["MVR All Trans", "PDF Maker", "PDF Play"]
    if role == "ADMIN":
        return base 
    elif role == "QA":
        return base
    elif role == "MAKER":
        return []
    return []

with st.sidebar:
    st.markdown(f"### 👋 Welcome, **{st.session_state['username']}**")
    st.markdown(f"**Role:** {st.session_state['role']}")
    st.markdown("---")

    menu_options = get_menu_options(st.session_state["role"])
    if menu_options:
        menu = st.radio("📋 Menu", menu_options, label_visibility="collapsed")
    else:
        st.warning("No menu options available.")
        menu = None

    st.markdown("---")
    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()
    st.caption("Built with Yogaraj ")

if menu == "MVR All Trans":
    from Alltran import Alltrans
    
    st.title("Alltrans Process")

    main_file = st.file_uploader("Upload MVR File", type=["xlsx"])
    lookup_file = st.file_uploader("Upload Client Excel", type=["xlsx","CSV"])

    if main_file and lookup_file:
        if st.button("Process"):
            try:
                main_bytes = main_file.read()
                lookup_bytes = lookup_file.read()

            
                from openpyxl import load_workbook
                lookup_wb = load_workbook(io.BytesIO(lookup_bytes), read_only=True, data_only=True)
                sheets = lookup_wb.sheetnames
                chosen_sheet = None
                if len(sheets) > 1:
                    chosen_sheet = st.selectbox("Lookup workbook has multiple sheets. Choose one:", options=sheets)
                else:
                    chosen_sheet = sheets[0]
                    st.write(f"From Client Excel: {chosen_sheet}")

                gen = Alltrans(template_path="Template.xlsx",
                            alltrans_sheet="All Trans",
                            alltrans_header_row=4,#Fixed constand row for Column Headers
                            mvr_sheet_name="MVR")
                out = gen.run(main_bytes, lookup_bytes, chosen_lookup_sheet=chosen_sheet, preview_rows=8)

                original_name = getattr(main_file, "name", None)
                base = original_name.rsplit(".", 1)[0] if original_name else "Final_Report"
                out_name = f"{base}.xlsx"
                st.success("Final report generated")
                st.download_button("Download Final Report", data=out, file_name=out_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Error: {e}")


elif menu == "PDF Maker":
    uploaded = st.file_uploader("Upload PDF", type=["pdf"])

    if uploaded:
        temp_input = "uploaded.pdf"
        with open(temp_input, "wb") as f:
            f.write(uploaded.read())

        if st.button("Process PDF"):
            st.info("Processing... bruhh...")

            output_file = "flattened.pdf"
            process_pdf(temp_input, output_file)

            with open(output_file, "rb") as f:
                st.download_button(
                    "Download Processed PDF",
                    f,
                    file_name="flattened.pdf",
                    mime="application/pdf"
                )

            # cleanup
            os.remove(temp_input)
            os.remove(output_file)

elif menu == "PDF Play":
    from pdf_play import WordToPDF, ExcelToPDF, ImageToPDF, PDFCompressor, PDFMerger, PDFSplitter
    st.title("PDF Play Ground")

    tool_option = st.selectbox("Select Tool", [
        "Word → PDF", 
        "Excel → PDF", 
        "Images → PDF", 
        "Merge PDFs",
        "Split PDF",
        "Compress PDF"
    ])

    # Helper for saving uploaded file
    def save_uploaded(uploaded):
        suffix = os.path.splitext(uploaded.name)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded.getbuffer())
            return tmp.name

    if tool_option == "Word → PDF":
        st.subheader("Word (.doc/.docx) to PDF")
        uploaded = st.file_uploader("Upload Word Document", type=["doc", "docx"])
        if uploaded and st.button("Convert"):
            with st.spinner("Converting Word to PDF..."):
                tmp_path = save_uploaded(uploaded)
                try:
                    converter = WordToPDF()
                    out_path = converter.convert(tmp_path)
                    
                    with open(out_path, "rb") as f:
                        st.download_button("Download PDF", f, file_name=f"{uploaded.name}.pdf", mime="application/pdf")
                    os.remove(out_path)
                except Exception as e:
                    st.error(f"Error: {e}")
                finally:
                    os.remove(tmp_path)

    elif tool_option == "Excel → PDF":
        st.subheader("Excel (.xls/.xlsx) to PDF")
        uploaded = st.file_uploader("Upload Excel File", type=["xls", "xlsx"])
        if uploaded and st.button("Convert"):
            with st.spinner("Converting Excel to PDF..."):
                tmp_path = save_uploaded(uploaded)
                try:
                    converter = ExcelToPDF()
                    out_path = converter.convert(tmp_path)
                    
                    with open(out_path, "rb") as f:
                        st.download_button("Download PDF", f, file_name=f"{uploaded.name}.pdf", mime="application/pdf")
                    os.remove(out_path)
                except Exception as e:
                    st.error(f"Error: {e}")
                finally:
                    os.remove(tmp_path)

    elif tool_option == "Images → PDF":
        st.subheader("Images to PDF")
        uploaded_files = st.file_uploader("Upload Images", type=["jpg", "jpeg", "png"], accept_multiple_files=True)
        if uploaded_files and st.button("Convert"):
            with st.spinner("Converting Images to PDF..."):
                img_paths = []
                # Save all images
                for up in uploaded_files:
                    img_paths.append(save_uploaded(up))
                
                try:
                    converter = ImageToPDF()
                    out_path = converter.convert(img_paths)
                    
                    with open(out_path, "rb") as f:
                        st.download_button("Download PDF", f, file_name="images.pdf", mime="application/pdf")
                    # Cleanup output
                    if os.path.exists(out_path):
                        os.remove(out_path)
                except Exception as e:
                    st.error(f"Error: {e}")
                finally:
                    # Cleanup inputs
                    for p in img_paths:
                        if os.path.exists(p):
                            os.remove(p)

    #             tmp_path = save_uploaded(uploaded)
    #             try:
    #                 processor = PDFOCR()
    #                 out_path = processor.process(tmp_path)
                    
    #                 with open(out_path, "rb") as f:
    #                     st.download_button("Download Searchable PDF", f, file_name=f"{uploaded.name}_ocr.pdf", mime="application/pdf")
    #                 os.remove(out_path)
    #             except Exception as e:
    #                 st.error(f"Error: {e}")
    #             finally:
    #                 os.remove(tmp_path)

    elif tool_option == "Compress PDF":
        st.subheader("Compress PDF File Size")
        uploaded = st.file_uploader("Upload PDF", type=["pdf"])
        level = st.select_slider("Compression Level", options=["Low", "Medium", "High"], value="Medium")
        
        if uploaded and st.button("Compress"):
            with st.spinner("Compressing PDF..."):
                tmp_path = save_uploaded(uploaded)
                try:
                    compressor = PDFCompressor()
                    out_path = compressor.compress(tmp_path, level.lower())
                    
                    # Show stats
                    original_size = os.path.getsize(tmp_path)
                    new_size = os.path.getsize(out_path)
                    reduction = (1 - new_size/original_size) * 100
                    st.success(f"Reduced by {reduction:.1f}% ({original_size/(1024*1024):.2f}MB → {new_size/(1024*1024):.2f}MB)")
                    
                    with open(out_path, "rb") as f:
                        st.download_button("Download Compressed PDF", f, file_name=f"compressed_{uploaded.name}", mime="application/pdf")
                    os.remove(out_path)
                except Exception as e:
                    st.error(f"Error: {e}")
                finally:
                    os.remove(tmp_path)

    elif tool_option == "Merge PDFs":
        st.subheader("Merge Multiple PDFs")
        uploaded_files = st.file_uploader("Upload PDF files", type=["pdf"], accept_multiple_files=True)
        
        # --- State Management for Order ---
        if "merge_order" not in st.session_state:
            st.session_state.merge_order = []
            
        # Helper to identify files uniquely
        def get_file_id(file):
            return f"{file.name}_{file.size}"

        # Reconcile uploaded files with session order
        if uploaded_files:
            current_files_map = {get_file_id(f): f for f in uploaded_files}
            current_ids = set(current_files_map.keys())
            
            # 1. Remove files that were un-uploaded
            st.session_state.merge_order = [fid for fid in st.session_state.merge_order if fid in current_ids]
            
            # 2. Add new files to the end
            for fid in current_files_map:
                if fid not in st.session_state.merge_order:
                    st.session_state.merge_order.append(fid)
                    
            # --- Reordering UI ---
            st.write("### 🔢 Organize Files")
            st.caption("Reorder files using Up/Down buttons. The final PDF will follow this order.")
            
            # We need callbacks for buttons
            def move_up(idx):
                if idx > 0:
                    st.session_state.merge_order[idx], st.session_state.merge_order[idx-1] = \
                    st.session_state.merge_order[idx-1], st.session_state.merge_order[idx]
            
            def move_down(idx):
                if idx < len(st.session_state.merge_order) - 1:
                    st.session_state.merge_order[idx], st.session_state.merge_order[idx+1] = \
                    st.session_state.merge_order[idx+1], st.session_state.merge_order[idx]

            # Display list with buttons
            for i, fid in enumerate(st.session_state.merge_order):
                f_obj = current_files_map[fid]
                c1, c2, c3 = st.columns([6, 1, 1])
                with c1:
                    st.text(f"{i+1}. {f_obj.name}")
                with c2:
                    st.button("⬆", key=f"up_{fid}", on_click=move_up, args=(i,), disabled=(i==0))
                with c3:
                    st.button("⬇", key=f"down_{fid}", on_click=move_down, args=(i,), disabled=(i==len(st.session_state.merge_order)-1))

            # --- Merge Logic ---
            if len(st.session_state.merge_order) >= 2:
                st.markdown("---")
                
                # Output Filename Logic
                default_name = current_files_map[st.session_state.merge_order[0]].name.rsplit('.', 1)[0] + "_merged"
                output_name_input = st.text_input("Output Filename (without .pdf)", value=default_name)
                final_output_name = f"{output_name_input}.pdf" if not output_name_input.lower().endswith('.pdf') else output_name_input

                if st.button("Merge PDFs Now"):
                    with st.spinner("Merging PDFs in specific order..."):
                        pdf_paths = []
                        temp_path_map = {} # Keep track to delete later
                        
                        try:
                            # Save files IN ORDER
                            for fid in st.session_state.merge_order:
                                f_obj = current_files_map[fid]
                                path = save_uploaded(f_obj)
                                pdf_paths.append(path)
                                temp_path_map[path] = True
                            
                            merger = PDFMerger()
                            # Pass strict ordered list
                            out_path = merger.merge(pdf_paths)
                            
                            with open(out_path, "rb") as f:
                                st.download_button(
                                    label=f"⬇ Download {final_output_name}", 
                                    data=f, 
                                    file_name=final_output_name, 
                                    mime="application/pdf"
                                )
                            
                            # Cleanup output
                            if os.path.exists(out_path):
                                os.remove(out_path)
                                
                        except Exception as e:
                            st.error(f"Merge Error: {e}")
                            
                        finally:
                            # Cleanup inputs
                            for p in pdf_paths:
                                if os.path.exists(p):
                                    try:
                                        os.remove(p)
                                    except:
                                        pass
            else:
                 st.info("Upload at least 2 files to enable merging.")
        
        else:
            # Clear state if no files uploaded
            st.session_state.merge_order = []
            st.warning("Please upload at least 2 PDF files")

    elif tool_option == "Split PDF":
        st.subheader("Split PDF / Extract Pages")
        uploaded = st.file_uploader("Upload PDF to Split", type=["pdf"])
        
        if uploaded:
            # Show basic info
            try:
                import fitz
                with fitz.open(stream=uploaded.read(), filetype="pdf") as doc:
                    num_pages = len(doc)
                uploaded.seek(0) # Reset pointer
                st.info(f"Total Pages: {num_pages}")
                
                page_range_str = st.text_input("Enter Page Numbers to Extract (e.g., 1, 3-5, 8)", help="Comma-separated numbers or ranges. Example: 1, 3-5 extracts pages 1, 3, 4, 5.")
                
                if st.button("Extract Pages"):
                    if not page_range_str.strip():
                        st.error("Please enter a page range.")
                    else:
                        with st.spinner("Extracting pages..."):
                            tmp_path = save_uploaded(uploaded)
                            try:
                                # Parse Range String to List of Integers (0-based)
                                selected_pages = []
                                parts = page_range_str.split(',')
                                for part in parts:
                                    part = part.strip()
                                    if '-' in part:
                                        start, end = part.split('-')
                                        start, end = int(start), int(end)
                                        # Handle Py vs Human indexing (Human 1-based -> Py 0-based)
                                        # inclusive range
                                        selected_pages.extend(range(start-1, end))
                                    else:
                                        selected_pages.append(int(part)-1)
                                
                                # Remove duplicates and sort
                                selected_pages = sorted(list(set(selected_pages)))
                                
                                splitter = PDFSplitter()
                                out_path = splitter.extract_pages(tmp_path, selected_pages)
                                
                                output_name = f"{os.path.splitext(uploaded.name)[0]}_extracted.pdf"
                                with open(out_path, "rb") as f:
                                    st.download_button("Download Extracted PDF", f, file_name=output_name, mime="application/pdf")
                                    
                                os.remove(out_path)
                                
                            except ValueError as ve:
                                st.error(f"Invalid input: {ve}")
                            except Exception as e:
                                st.error(f"Error: {e}")
                            finally:
                                if os.path.exists(tmp_path):
                                    os.remove(tmp_path)
            except Exception as e:
                st.error(f"Error reading file info: {e}")