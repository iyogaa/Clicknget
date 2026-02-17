import os
import fitz  # pymupdf
from PIL import Image

# Try to import optional dependencies
try:
    import mammoth
except ImportError:
    mammoth = None

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfgen import canvas


class WordToPDF:
    def __init__(self):
        if not mammoth:
            raise ImportError("mammoth library is required for Word conversion.")

    def convert(self, input_path):
        """Convert Word docx to PDF by extracting text."""
        output_path = input_path.rsplit('.', 1)[0] + ".pdf"
        
        with open(input_path, "rb") as docx_file:
            result = mammoth.extract_raw_text(docx_file)
            text = result.value
            
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        for line in text.split('\n'):
            if line.strip():
                story.append(Paragraph(line, styles["Normal"]))
                story.append(Spacer(1, 6))
                
        doc.build(story)
        return output_path

class ExcelToPDF:
    def __init__(self):
        import openpyxl
        self.openpyxl = openpyxl

    def convert(self, input_path):
        """Convert Excel sheets to PDF using Manual Canvas Rendering for perfect spacing and no overlapping."""
        from openpyxl.utils import get_column_letter
        
        output_path = input_path.rsplit('.', 1)[0] + ".pdf"
        wb = self.openpyxl.load_workbook(input_path, data_only=True)
        
        c = canvas.Canvas(output_path, pagesize=A4)
        page_width, page_height = A4
        margin = 36 # 0.5 inch
        avail_w = page_width - (margin * 2)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Detect range
            max_row, max_col = 0, 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        max_row = max(max_row, cell.row)
                        max_col = max(max_col, cell.column)

            if max_row == 0 or max_col == 0:
                continue

            # 1. Dimensions & Scale
            orig_widths = []
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                w = ws.column_dimensions[col_letter].width or 8.43
                orig_widths.append(w * 7.0)

            content_w = sum(orig_widths)
            scale = avail_w / content_w if content_w > avail_w else 1.0
            if scale < 0.1: scale = 0.1

            # 2. Render
            current_y = page_height - margin
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, current_y - 12, f"Sheet: {sheet_name}")
            current_y -= 35

            for r_idx in range(1, max_row + 1):
                row_h = (ws.row_dimensions[r_idx].height or 15) * scale
                
                if current_y - row_h < margin:
                    c.showPage()
                    current_y = page_height - margin
                
                current_x = margin
                
                for c_idx, ow in enumerate(orig_widths, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cw = ow * scale
                    
                    # Border
                    c.setStrokeColor(colors.lightgrey)
                    c.setLineWidth(0.5 * scale)
                    c.rect(current_x, current_y - row_h, cw, row_h, stroke=1, fill=0)
                    
                    if cell.value is not None:
                        text = str(cell.value).strip()
                        if text:
                            f_size = max(8 * scale, 6)
                            f_name = "Helvetica"
                            if cell.font and cell.font.bold: f_name = "Helvetica-Bold"
                            
                            c.setFont(f_name, f_size)
                            c.setFillColor(colors.black)
                            
                            padding = 2 * scale
                            tx = current_x + padding
                            ty = current_y - row_h + padding + (f_size * 0.15)
                            
                            # Overlap prevention: Truncate
                            atw = cw - (padding * 2)
                            if c.stringWidth(text, f_name, f_size) > atw:
                                td = text
                                while len(td) > 0 and c.stringWidth(td + "...", f_name, f_size) > atw:
                                    td = td[:-1]
                                td += "..."
                            else:
                                td = text

                            c.drawString(tx, ty, td)

                    current_x += cw
                
                current_y -= row_h

            c.showPage()

        wb.close()
        c.save()
        return output_path

class ImageToPDF:
    def __init__(self):
        pass

    def convert(self, image_files, output_filename="images.pdf"):
        if not image_files:
            raise ValueError("No images provided")
        
        images = []
        for img_path in image_files:
            img = Image.open(img_path)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            images.append(img)

        output_path = os.path.join(os.path.dirname(image_files[0]), output_filename)
        
        if len(images) == 1:
            images[0].save(output_path, "PDF", resolution=100.0)
        else:
            images[0].save(output_path, "PDF", save_all=True, append_images=images[1:], resolution=100.0)
            
        return output_path

class PDFCompressor:
    def __init__(self):
        pass

    def compress(self, input_path, level="medium"):
        output_path = input_path.rsplit('.', 1)[0] + "_compressed.pdf"
        doc = fitz.open(input_path)
        if level == "high":
            doc.save(output_path, garbage=4, deflate=True, clean=True)
        elif level == "medium":
            doc.save(output_path, garbage=3, deflate=True)
        else: # low
            doc.save(output_path, garbage=2, deflate=True)
        return output_path

class PDFMerger:
    def __init__(self):
        pass

    def merge(self, input_paths):
        if not input_paths:
            raise ValueError("No input files provided")
        merged = fitz.open()
        for path in input_paths:
            try:
                doc = fitz.open(path)
                merged.insert_pdf(doc)
            except Exception as e:
                raise Exception(f"Failed to merge {os.path.basename(path)}: {e}")
        output_path = input_paths[0].rsplit('.', 1)[0] + "_merged.pdf"
        merged.save(output_path)
        merged.close()
        return output_path

class PDFSplitter:
    def __init__(self):
        pass

    def extract_pages(self, input_path, page_numbers):
        doc = fitz.open(input_path)
        output_path = input_path.rsplit('.', 1)[0] + "_extracted.pdf"
        max_page = len(doc) - 1
        valid_pages = [p for p in page_numbers if 0 <= p <= max_page]
        if not valid_pages:
            raise ValueError("No valid pages selected.")
        doc.select(valid_pages)
        doc.save(output_path)
        doc.close()
        return output_path
