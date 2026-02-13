import streamlit as st
import io
from Alltran import Alltrans
from openpyxl import load_workbook

def run_mvr_all_trans():
    st.title("Alltrans Process")

    main_file = st.file_uploader("Upload MVR File", type=["xlsx", "xls", "csv"])
    lookup_file = st.file_uploader("Upload Client Excel", type=["xlsx", "xls", "csv"])

    if main_file and lookup_file:
        if st.button("Process"):
            try:
                main_bytes = main_file.read()
                lookup_bytes = lookup_file.read()

                chosen_sheet = None
                try:
                    # Try to see if it has sheets (Excel)
                    lookup_wb = load_workbook(io.BytesIO(lookup_bytes), read_only=True)
                    sheets = lookup_wb.sheetnames
                    if len(sheets) > 1:
                        chosen_sheet = st.selectbox("Lookup workbook has multiple sheets. Choose one:", options=sheets)
                    else:
                        chosen_sheet = sheets[0]
                        st.write(f"Sheet detected: {chosen_sheet}")
                except Exception:
                    # Likely a CSV or not a standard Excel file
                    st.write("File format: CSV/Plain Text (No sheets)")
                    chosen_sheet = None

                gen = Alltrans(template_path="Template.xlsx",
                            alltrans_sheet="All Trans",
                            alltrans_header_row=4,#Fixed constant row for Column Headers
                            mvr_sheet_name="MVR")
                
                with st.spinner("Processing Alltrans..."):
                    out = gen.run(main_bytes, lookup_bytes, chosen_lookup_sheet=chosen_sheet, preview_rows=8)

                original_name = getattr(main_file, "name", None)
                base = original_name.rsplit(".", 1)[0] if original_name else "Final_Report"
                out_name = f"{base}.xlsx"
                st.success("Final report generated")
                st.download_button("Download Final Report", data=out, file_name=out_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback
                st.code(traceback.format_exc())
