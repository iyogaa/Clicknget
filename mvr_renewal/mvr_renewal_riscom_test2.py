from datetime import datetime
import string
from io import BytesIO
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from dateutil import parser
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import PieChart3D, BarChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    RichTextProperties,
    ParagraphProperties,
    Paragraph,
    CharacterProperties,
)
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1, FORMAT_NUMBER
import re
from openpyxl.cell import MergedCell
import io
import zipfile
from dateutil.relativedelta import relativedelta
import os
import sys
import traceback
from fuzzywuzzy import fuzz

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__)))

if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)


    

DATE_FORMAT_STRING = "%m/%d/%Y"

def unpack_tuples(cell_value):
    if isinstance(cell_value, tuple):
        return cell_value[-1]
    return cell_value

first_sheet = True
blue_header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
blue_header_fill = PatternFill(
    start_color="0033CCCC", end_color="0033CCCC", fill_type="solid"
)


class Chart:
    def __init__(self, title="", data_length=5, chart_type=None, min_col_label=None, max_col_label=None, 
                 min_col_data=None, max_col_data=None, min_row_label=None, min_row_data=None, 
                 max_row_label=None, max_row_data=None, excel_column=None, excel_row=None, 
                 orient="col", dimensions=2):
        self.title = title
        self.data_length = data_length
        self.chart_type = chart_type
        self.min_col_label = min_col_label
        self.max_col_label = max_col_label
        self.min_col_data = min_col_data
        self.max_col_data = max_col_data
        self.min_row_label = min_row_label
        self.min_row_data = min_row_data
        self.max_row_label = max_row_label
        self.max_row_data = max_row_data
        self.excel_column = excel_column
        self.excel_row = excel_row
        self.orient = orient
        self.dimensions = dimensions


def attach_pie_chart(worksheet, chart_obj: Chart):
    datalabel = DataLabelList()
    datalabel.showPercent = True
    # datalabel.showVal = True
    pie = PieChart3D(dLbls=datalabel)
    labels = Reference(
        worksheet,
        min_col=chart_obj.min_col_label,
        min_row=chart_obj.min_row_label,
        max_row=chart_obj.data_length + 2,
    )
    data = Reference(
        worksheet,
        min_col=chart_obj.min_col_data,
        min_row=chart_obj.min_row_data,
        max_row=chart_obj.data_length + 2,
        max_col=chart_obj.max_col_data,
    )
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    pie.title = chart_obj.title

    # pie.legend.position = None
    # pie.legend.overlay = '0'

    worksheet.add_chart(pie, f"{chart_obj.excel_column}{chart_obj.excel_row}")


def attach_bar_chart(worksheet, chart_obj: Chart, chart_first):
    if chart_first:
        start_row = 30
    else:
        start_row = 2
    axis = CharacterProperties(sz=1000)
    rot = (
        RichTextProperties(vert="vert270")
        if chart_obj.orient == "col"
        else RichTextProperties()
    )
    datalabel = DataLabelList()
    datalabel.showPercent = True
    datalabel.showVal = True
    datalabel.textProperties = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot
    )

    chart_size = chart_obj.data_length * 3
    bar_chart = (
        BarChart3D(dLbls=datalabel)
        if chart_obj.dimensions == 3
        else BarChart(dLbls=datalabel)
    )
    bar_chart.height, bar_chart.width = (
        (bar_chart.height, max(5 + chart_size, bar_chart.width))
        if chart_obj.orient == "col"
        else (max(chart_size, bar_chart.height), bar_chart.width)
    )
    bar_chart.type = chart_obj.orient
    labels = Reference(
        worksheet,
        min_col=chart_obj.min_col_label,
        min_row=chart_obj.min_row_label,
        max_row=chart_obj.data_length + start_row,
        max_col=(
            chart_obj.max_col_label
            if chart_obj.max_col_label
            else chart_obj.min_col_label
        ),
    )
    data = Reference(
        worksheet,
        min_col=chart_obj.min_col_data,
        min_row=chart_obj.min_row_data,
        max_row=chart_obj.data_length + start_row,
        max_col=chart_obj.max_col_data,
    )
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(labels)

    # for series in bar_chart.series:
    #     for point in series.points:
    #         point.dLbls.showVal = True

    bar_chart.title = chart_obj.title
    worksheet.add_chart(bar_chart, f"{chart_obj.excel_column}{chart_obj.excel_row}")





def apply_formatting_new(worksheet, min_col, max_col):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(min_row=29, min_col=min_col, max_col=max_col):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 30:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row > 2 and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            cell.number_format = FORMAT_NUMBER
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def apply_formatting(worksheet, min_col, max_col, custom_format_col_list=None, custom_format_cell_list=None):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(min_col=min_col, max_col=max_col):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 2:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row > 2 and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            # if column or the cell from summary sheet is to be kept integer
            if (custom_format_col_list and not isinstance(cell,
                                                          MergedCell) and cell.column_letter in custom_format_col_list) or (
                    custom_format_cell_list and not isinstance(cell,
                                                               MergedCell) and cell.coordinate in custom_format_cell_list):
                cell.number_format = FORMAT_NUMBER
            else:
                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def add_blue_header(sheet, header_text, chart_first=False, last_column=5, start_row=None):
    if not start_row:
        sheet.insert_rows(1)
    elif start_row and start_row == 2:
        sheet.insert_rows(1)
    else:
        sheet.insert_rows(start_row - 2)

    if chart_first:
        start_row = 29
    if start_row:
        start_row = start_row - 1
    else:
        start_row = start_row if start_row else 1
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_column)

    header_cell = sheet.cell(row=start_row, column=1)
    header_cell.value = header_text
    header_cell.font = blue_header_font  # Font with smaller size
    header_cell.fill = blue_header_fill
    header_cell.alignment = Alignment(horizontal="center")


def strip_non_numeric_and_leading_zeros(input_string):
    numeric_part = re.sub(r'\D', '', str(input_string))
    return numeric_part.lstrip('0')


def match_driver_names(name_driver, name_grouped):
    set_ratio = fuzz.token_set_ratio(name_driver, name_grouped)
    sort_ratio = fuzz.token_sort_ratio(name_driver, name_grouped)
    return (set_ratio + sort_ratio) / 2


def filter_drivers_rsc(driver_row, grouped_row):
    name_driver = driver_row['Driver First Name'] + " " + driver_row['Driver Last Name']
    name_grouped = grouped_row['Driver First Name'] + " " + grouped_row['Driver Last Name']
    cdl_driver, cdl_grouped = driver_row['License Number'], grouped_row['License Number']
    dob_driver, dob_grouped = driver_row['Driver Date of Birth'], grouped_row['Driver Date of Birth']

    if strip_non_numeric_and_leading_zeros(cdl_driver) == strip_non_numeric_and_leading_zeros(cdl_grouped):
        return True

    try:
        dob_driver_parsed = parser.parse(dob_driver).date()
        dob_grouped_parsed = parser.parse(dob_grouped).date()
    except (ValueError, TypeError):
        dob_driver_parsed = None
        dob_grouped_parsed = None

    if dob_driver_parsed and dob_grouped_parsed:
        if dob_driver_parsed == dob_grouped_parsed:
            return match_driver_names(name_driver, name_grouped) >= 80
    else:
        return match_driver_names(name_driver, name_grouped) >= 80

    return False


def split_driver_name(full_name):
    SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}
    if pd.isnull(full_name) or full_name.strip() == "":
        return pd.Series(["", ""])
    full_name = full_name.strip()
    if ',' in full_name:
        # Split on the first comma
        last, first = [p.strip() for p in full_name.split(',', 1)]
        first_parts = first.split()
        # Check if the last part is a suffix
        if first_parts and first_parts[-1].upper() in SUFFIXES:
            suffix = first_parts.pop(-1)
            last = f"{last} {suffix}"
        first_name = " ".join(first_parts)
        last_name = last
    else:
        # No comma: check for suffix at the end
        parts = full_name.split()
        if len(parts) >= 3 and parts[-1].upper() in SUFFIXES:
            # e.g., "Mario Espinola II"
            last_name = f"{parts[0]} {parts[-1]}"
            first_name = " ".join(parts[1:-1])
        elif len(parts) >= 2:
            # e.g., "Smith John"
            last_name = parts[0]
            first_name = " ".join(parts[1:])
        else:
            first_name = parts[0]
            last_name = ""
    return pd.Series([first_name, last_name])


def calculate_age(dob_str):
    if dob_str:
        try:
            dob = parser.parse(dob_str).date()
            today = datetime.now().date()
            # More precise: subtract one if birthday hasn't occurred yet this year
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
            return age
        except Exception:
            return ""
    return ""


def apply_formatting_mvr_owned(
        worksheet, min_col, max_col, start_row=None, end_row=None
):
    # Define styles
    font_with_size_10 = Font(bold=True, size=10)
    font_with_white_color = Font(color="FFFFFF")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    light_grey_fill = PatternFill(
        start_color="E5E4E2", end_color="E5E4E2", fill_type="solid"
    )
    # white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White fill
    black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    white_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    # Loop through all rows in the worksheet
    for row in worksheet.iter_rows(
            min_col=min_col, max_col=max_col, min_row=start_row - 1, max_row=end_row
    ):
        contains_data = any(cell.value for cell in row)
        for cell in row:
            # Apply grey fill only to the second row
            if cell.row == 2:
                cell.fill = grey_fill
                cell.font = font_with_white_color
            if cell.row >= start_row and cell.row <= end_row and cell.row % 2 == 0:
                cell.fill = light_grey_fill
            cell.number_format = "0"
            # Apply font size 10 and black border if the row contains data
            if contains_data:
                cell.font = font_with_size_10
                cell.border = black_border
                # apply text-wrap alignment to cell data, beyond Description row
                # if cell.row > 1:
                #     cell.alignment = Alignment(wrap_text=True)
            # else:
            #     cell.border = white_border
            #     cell.fill = white_fill


def write_excel_sheet_mvr_owned(
        data1, workbook, sheet_name, header_text, _first_sheet, start_row=2, skip_blue_header=False
):
    # global _first_sheet
    # charts = data[1]
    start_row_real = start_row
    data = data1
    if isinstance(data, dict):
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # creating tables for the graph data
        add_blue_header(worksheet, header_text, last_column=2, start_row=start_row_real)
        # if start_row != 2:
        # start_row = start_row-1
        start_row = start_row
        start_column = 1
        for row in data.items():
            for col_index, value in enumerate(row, start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        apply_formatting_mvr_owned(
            worksheet,
            min_col=1,
            max_col=2,
            start_row=start_row_real,
            end_row=start_row_real + len(data) - 1,
        )

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].auto_size = True

    elif isinstance(data, list) and data: # Added check for empty list
        if _first_sheet:
            worksheet = workbook.active
            worksheet.title = sheet_name
            _first_sheet = False
        elif sheet_name in [sheet.title for sheet in workbook._sheets]:
            worksheet = workbook[sheet_name]
            workbook.active = worksheet
        else:
            worksheet = workbook.create_sheet(sheet_name)
        start_row = start_row - 1
        start_column = 1
        last_column = 5

        header_row_flag = True
        for row in data:
            last_column = len(row)
            if header_row_flag:
                header_row_flag = False
                # Write header
                for col_index, key in enumerate(row.keys(), start=start_column):
                    worksheet.cell(
                        row=start_row,
                        column=col_index,
                        value=key
                    )
                start_row += 1
            # Write data row
            for col_index, value in enumerate(row.values(), start=start_column):
                worksheet.cell(row=start_row, column=col_index, value=value)
            start_row += 1
        
        if not skip_blue_header:
            add_blue_header(worksheet, header_text, last_column, start_row=start_row_real)
        else:
            start_row_real -= 1  # avoid one extra grey row
        
        if data: # Only apply formatting if data exists
            apply_formatting_mvr_owned(
                worksheet,
                min_col=1,
                max_col=len(data[0]),
                start_row=start_row_real,
                end_row=start_row_real + len(data),
            )


def custom_major_minor_3_5(x, violation_dates, three_years_ago, five_years_ago):
    values = [str(value).strip().lower() for value in x]
    # Check for major violations in last 5 years
    recent_5 = [(v, d) for v, d in zip(values, violation_dates) if pd.notnull(d) and d >= five_years_ago]
    recent_5_values = [v for v, d in recent_5]
    if "major" in recent_5_values:
        return "Major"
    # Check for minor violations in last 3 years
    recent_3 = [(v, d) for v, d in zip(values, violation_dates) if pd.notnull(d) and d >= three_years_ago]
    recent_3_values = [v for v, d in recent_3]
    if "minor" in recent_3_values:
        return "Minor"
    return ""


def custom_concat(x) -> str:
    unique_values = set(x)
    non_empty_values = [value for value in unique_values if value != ""]
    res = ", ".join(map(str, non_empty_values))
    return res


def concat_without_clean_mvr(existing, to_add):
    """
    Concatenates two strings, removing 'Clean MVR' (case-insensitive) and '_x000d_\' artifacts from both.
    Adds a comma only if both are non-empty.
    If existing has 'Clean MVR' and to_add is empty, return existing as is.
    If existing has 'Clean MVR' and to_add has other values, remove 'Clean MVR' from existing.
    Also removes '_x000d_\' artifacts from both strings.
    """

    def clean_string(s):
        # Remove 'Clean MVR' (case-insensitive)
        # s = re.sub(r'(?i)\bclean mvr\b', '', str(s))
        # Remove '_x000d_\' (case-insensitive, with optional trailing whitespace/newline)
        s = re.sub(r'(?i)_x000d_\\[\s\n]*', '', s)
        # Strip leading/trailing commas and whitespace
        return s.strip(', ').strip()

    cleaned_to_add = clean_string(to_add)
    if cleaned_to_add:
        cleaned_existing = clean_string(existing)
        if cleaned_existing:
            return f"{cleaned_existing},{cleaned_to_add}"
        else:
            return cleaned_to_add
    else:
        return clean_string(existing)


def status_and_comments(row, has_heavy_vehicle):
    comments = []
    status = ""

    # Parse age
    try:
        age_val = row.get("Age")
        age = int(age_val) if age_val not in [None, ""] else 0
    except Exception:
        age = 0

    # Parse violation counts (already pre-calculated for correct timeframes: Minor 3yr, Major 5yr)
    try:
        minor_count = int(row.get("Minor Count", 0) or 0)
    except Exception:
        minor_count = 0
    try:
        major_count = int(row.get("Major Count", 0) or 0)
    except Exception:
        major_count = 0

    # Medical status & expiration
    med_status = str(row.get("Medical Status", "")).strip().lower()
    med_exp = row.get("Medical Expiration Date", "")
    
    # Restrictions
    restrictions = str(row.get("Restrictions", "")).strip()

    # Violation Descriptions
    violation_desc = row.get("Violation Description", "")

    # License Status
    license_status_raw = str(row.get("License Status", "")).strip()
    license_status_lower = license_status_raw.lower()

    # --- Pending Logic ---
    pending = False

    # 1. Age Rules
    if age < 21:
        pending = True
        if "Age Requirement" not in comments:
            comments.append("Age Requirement")
    if age > 69:
        pending = True
        if "Age Requirement" not in comments:
            comments.append("Age Requirement")

    # 2. Violation Rules
    # Minor > 3 (last 3 years) -> Pending
    # Major > 0 (last 5 years) -> Pending
    if minor_count > 3 or major_count > 0:
        pending = True
        # Add ALL violations comma-separated in Comments
        if violation_desc:
            comments.append(violation_desc)

    # 3. Medical Rules
    if med_status == "not certified":
        pending = True
        comments.append("Medical Status Not Certified")
    
    if med_exp:
        try:
            if isinstance(med_exp, datetime):
                med_exp_date = med_exp
            else:
                med_exp_date = parser.parse(str(med_exp))
            
            # Compare with current date (ignoring time)
            if med_exp_date.date() < datetime.now().date():
                pending = True
                comments.append("Medical Expired")
        except Exception:
            pass

    # 4. MVR Presence Rule is handled in post-processing logic usually
    
    # 5. Restrictions
    if restrictions:
        allowed_restrictions = {
            'glasses/contacts',
            'with corrective lenses',
            'corrective lenses',
            'corrective lens',
            'b - corrective lenses',
            'wear corrective lenses',
            'glasses, contact lenses',
            'motorcycle',
            'corrective lenses(glasses/contacts)'
        }
        
        
        restriction_items = [r.strip() for r in restrictions.split(',') if r.strip()]
        
        extra_restrictions = []
        for r in restriction_items:
            if r.lower() not in allowed_restrictions:
                extra_restrictions.append(r)

        if extra_restrictions:
            pending = True
            comments.append(", ".join(extra_restrictions))
        
            

    # 6. License Status
    if not license_status_raw:
        comments.append("License status unknown")
        pending = True
        
    if "surrender" in license_status_lower:
         comments.append("License Surrender")
         pending = True
         
    if "cancelled" in license_status_lower:
        comments.append("License Cancelled")
        pending = True

    if "suspended" in license_status_lower or "blocked" in license_status_lower:
        comments.append("License Suspended")
        pending = True

    # Final Status Determination
    if pending:
        status = "Pending"
    else:
        # Check Approved conditions
        # If Minor <= 3 AND Major == 0 -> Status = Approved
        
        if minor_count <= 3 and major_count == 0:
             status = "Approved"
             if not comments:
                 comments.append("Clean MVR")
        else:
             status = "Pending" # Fallback

    # Clean up comments
    comments = [c for c in comments if c]
    comments_str = ", ".join(comments)
    
    return pd.Series({"Status": status, "Comments": comments_str})



def filter_and_cleanup_mvr_data(mvr_data):
    """
    Filters out rows where exp_drivers_list_status is 'deleted' and mvr_list_flag is False,
    then drops the flag/status columns from the DataFrame.
    Returns a list of dicts.
    """
    # Convert to DataFrame
    df = pd.DataFrame(mvr_data)

    # Drop the specified columns
    cols_to_drop = [
        "mvr_list_flag",
        "driver_list_flag",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count 1 Year",
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year"
    ]
    df = df.drop(columns=cols_to_drop, errors="ignore")

    # Return as list of dicts
    return df.to_dict(orient="records")




def match_by_license(driver_row, exp_row):
    cdl_driver = driver_row['License Number']
    cdl_exp = exp_row['License Number']
    return strip_non_numeric_and_leading_zeros(cdl_driver) == strip_non_numeric_and_leading_zeros(cdl_exp)

def match_by_name_dob(driver_row, exp_row):
    name_driver = driver_row['Driver First Name'] + " " + driver_row['Driver Last Name']
    name_exp = exp_row['Driver First Name'] + " " + exp_row['Driver Last Name']
    dob_driver = driver_row['Driver Date of Birth']
    dob_exp = exp_row['Driver Date of Birth']
    try:
        dob_driver_parsed = parser.parse(dob_driver).date()
        dob_exp_parsed = parser.parse(dob_exp).date()
    except (ValueError, TypeError):
        return False
    if dob_driver_parsed == dob_exp_parsed:
        return match_driver_names(name_driver, name_exp) >= 80
    return False

def superset_drivers(drivers_df):
    """
    Prepares the drivers dataframe by adding necessary flags.
    Since exp_drivers_file is removed, this simply formats the drivers_df.
    """
    result_df = drivers_df.copy()
    result_df['driver_list_flag'] = True
    result_df['mvr_list_flag'] = False
    
    # Initialize columns expected by downstream logic
    columns_to_init = [
        'Comments'
    ]
    for col in columns_to_init:
        if col not in result_df.columns:
            result_df[col] = ""
            
    return result_df


def postprocess_mvr_data(mvr_data):
    """
    Modifies mvr_data in place according to MVR vs Driver Application logic.
    """
    for row in mvr_data:
        drv_flag = bool(row.get("driver_list_flag"))
        mvr_flag = bool(row.get("mvr_list_flag"))
        
        # Rule 7: MVR Received But Driver Not in Application
        if mvr_flag and not drv_flag:
            row["Status"] = "Pending"
            row["Comments"] = "Mvr received - driver not shown on application"
            
        # Rule 4: Eligible driver MVR not present
        elif drv_flag and not mvr_flag:
            row["Status"] = "Pending"
            row["Comments"] = "MVR Needed"
            row["MVR Received"] = "FALSE"
        
        # Normal Case: Driver in App AND MVR Received
        elif drv_flag and mvr_flag:
            row["MVR Received"] = "TRUE"
            # Status and comments are already calculated in status_and_comments based on the MVR data.
            # However, we need to ensure "MVR Needed" is NOT added.
            # And logic from status_and_comments persists.
            pass
            
    return mvr_data

def fill_missing_dob(df, driver_df):
    """
    Fill missing Driver DOB values by matching license numbers with driver_df
    """
    if driver_df is None or driver_df.empty:
        return df
        
    # Extract DOB data from driver_df
    if 'License Number' in driver_df.columns and 'Driver Date of Birth' in driver_df.columns:
        driver_ref = driver_df[['License Number', 'Driver Date of Birth']].copy()
        driver_ref = driver_ref.dropna(subset=['License Number'])
        driver_ref['Driver Date of Birth'] = pd.to_datetime(driver_ref['Driver Date of Birth'], errors='coerce')
        
        # Remove duplicates
        driver_ref = driver_ref.drop_duplicates(subset=['License Number'], keep='first')
        driver_ref = driver_ref.dropna(subset=['Driver Date of Birth'])
        
        # Create mapping
        license_to_dob = dict(zip(driver_ref['License Number'], driver_ref['Driver Date of Birth']))
        
        # Fill missing DOB values in df
        mask_missing_dob = df["Driver DOB"].isna() | (df["Driver DOB"] == "") | df["Driver DOB"].isnull()
        
        for idx in df[mask_missing_dob].index:
            license_num = df.loc[idx, 'license_number']
            if pd.notna(license_num) and license_num in license_to_dob:
                df.loc[idx, "Driver DOB"] = license_to_dob[license_num]
                df.loc[idx, "driver_date_of_birth"] = license_to_dob[license_num]
                
    return df


def generate_mvr_data_sheet_for_drivers(df, driver_df, vehicle_df):
    df = df.copy()
    df = df[df['section_name'] == 'section_mvr']

    # Convert columns to datetime
    df['violation_date'] = pd.to_datetime(df['violation_date'], errors='coerce')
    df['violation_category_str'] = df['violation_category']
    df["Driver DOB"] = pd.to_datetime(df["driver_date_of_birth"], errors="coerce")
    df["license_expiration_date"] = pd.to_datetime(df["license_expiration_date"], errors="coerce")
    df["driver_hiring_date"] = pd.to_datetime(df["driver_hiring_date"], errors="coerce")
    df["mvr_generation_date"] = pd.to_datetime(df["mvr_generation_date"], errors="coerce")
    df["restrictions"] = df["restrictions"].fillna("")
    df['violation_category'] = df['violation_category'].fillna('')
    df['violation_description'] = df['violation_description'].fillna('')
    df['license_status'] = df['license_status'].fillna('')
    df['driver_first_name']= df['driver_first_name'].fillna('')
    df['driver_last_name']= df['driver_last_name'].fillna('')

    current_date = pd.Timestamp.now()
    one_year_ago = current_date - pd.Timedelta(days=365)
    three_years_ago = current_date - pd.DateOffset(years=3)
    five_years_ago = current_date - pd.DateOffset(years=5)
    df["age"] = df["Driver DOB"].apply(
        lambda dob: calculate_age(str(dob)) if pd.notnull(dob) else None
    )

    # Helper functions for aggregation
    def minor_count_last_three_years(x):
        # Comparison adjusted to ensure consistency with pandas datetime series
        mask = (x['violation_category'].str.lower() == 'minor') & (x['violation_date'] >= three_years_ago)
        return mask.sum()

    def minor_desc_last_three_years(x):
        mask = (x['violation_category'].str.lower() == 'minor') & (x['violation_date'] >= three_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def major_count_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= five_years_ago)
        return mask.sum()

    def major_desc_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= five_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def major_count_last_one_year(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= one_year_ago)
        return mask.sum()

    def major_desc_last_one_year(x):
        mask = (x['violation_category'].str.lower() == 'major') & (x['violation_date'] >= one_year_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def accident_count_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'accident') & (x['violation_date'] >= five_years_ago)
        return mask.sum()

    def accident_desc_last_five_years(x):
        mask = (x['violation_category'].str.lower() == 'accident') & (x['violation_date'] >= five_years_ago)
        return custom_concat(x.loc[mask, 'violation_description'])

    def custom_major_minor_group_3_5(x):
        return custom_major_minor_3_5(
            x['violation_category_str'],
            x['violation_date'],
            three_years_ago,
            five_years_ago
        )

    # Group and aggregate
    df = fill_missing_dob(df, driver_df)
    group_cols = ["driver_full_name", "license_number", "Driver DOB"]
    grouped = df.groupby(group_cols, sort=False)

    grouped_df = grouped.apply(lambda x: pd.Series({
        # "Driver DOB": x["Driver DOB"].iloc[0],
        "driver_first_name": x["driver_first_name"].iloc[0],
        "driver_last_name": x["driver_last_name"].iloc[0],
        "driver_hiring_date": x["driver_hiring_date"].iloc[0],
        "age": x["driver_date_of_birth"].apply(
            lambda dob: calculate_age(str(dob)) if pd.notnull(dob) else None).iloc[0],
        "mvr_generation_date": x["mvr_generation_date"].iloc[0],
        "license_state": x["license_state"].iloc[0],
        "license_class": x["license_class"].iloc[0],
        "license_status": x["license_status"].iloc[0],
        "license_expiration_date": x["license_expiration_date"].iloc[0],
        "medical_expiration_date": x["medical_expiration_date"].iloc[0],
        "medical_status": x["medical_status"].iloc[0],
        "restrictions": custom_concat(x["restrictions"]),
        "violation_category_str": custom_major_minor_group_3_5(x),
        "Minor Count": minor_count_last_three_years(x),
        "Minor Count 3 Year": minor_count_last_three_years(x),  # for post logic we will drop later
        "Minor Violation Description": minor_desc_last_three_years(x),
        "Minor Violation Description 3 Year": minor_desc_last_three_years(x),  # for post logic we will drop later
        "Major Count": major_count_last_five_years(x),
        "Major Violation Description": major_desc_last_five_years(x),
        "Major Count 1 Year": major_count_last_one_year(x),  # for post logic we will drop later
        "Major Violation Description 1 Year": major_desc_last_one_year(x),  # for post logic we will drop later
        "Major Violation Description 5 Year": major_desc_last_five_years(x),
        "Accident Count": accident_count_last_five_years(x),
        "Accident Violation Description": accident_desc_last_five_years(x),
    })).reset_index()

    # Combine all descriptions into a single column, skipping empty values
    desc_cols = [
        "Minor Violation Description",
        "Major Violation Description",
        "Accident Violation Description"
    ]
    grouped_df["violation_description"] = grouped_df[desc_cols].apply(
        lambda row: ", ".join([str(val) for val in row if val and str(val).strip() != ""]), axis=1
    )
    grouped_df = grouped_df.drop(columns=desc_cols)

    # Rename columns
    grouped_df.columns = [
        "Driver Full Name",
        "License Number",
        "Driver Date of Birth",
        # "Driver Date of Birth",
        "Driver First Name",
        "Driver Last Name",
        "Hire Date",
        "Age",
        "Date MVR Ordered",
        "License State",
        "License Type",
        "License Status",
        "License Expiration Date",
        "Medical Expiration Date",
        "Medical Status",
        "Restrictions",
        "Violation Category",
        "Minor Count",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count",
        "Major Count 1 Year",
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year",
        "Accident Count",
        "Violation Description"
    ]
    # Calculate initial Status and Comments based on MVR data

    # Convert relevant columns to integers (handle NaN values)
    grouped_df["Accident Count"] = (
        pd.to_numeric(grouped_df["Accident Count"], errors="coerce")
        .apply(lambda x: int(x) if pd.notna(x) else x)
    )

    grouped_df["Minor Count"] = (
        pd.to_numeric(grouped_df["Minor Count"], errors="coerce").apply(lambda x: int(x) if pd.notna(x) else x)
    )
    grouped_df["Major Count"] = (
        pd.to_numeric(grouped_df["Major Count"], errors="coerce").apply(lambda x: int(x) if pd.notna(x) else x)
    )

    grouped_df["Total Incidents"] = (
        grouped_df[["Accident Count", "Minor Count", "Major Count"]]
        .sum(axis=1)
        .apply(lambda x: pd.NA if x == 0 else int(x) if pd.notna(x) else x)
    )
    grouped_df[["Minor Count", "Major Count", "Accident Count"]] = grouped_df[
        ["Minor Count", "Major Count", "Accident Count"]
    ].replace(0, "")

    # Format 'Driver Date of Birth' and 'License Expiration Date' to 'mm/dd/yyyy'
    grouped_df["Driver Date of Birth"] = grouped_df["Driver Date of Birth"].apply(
        lambda dob: dob.strftime("%m/%d/%Y") if pd.notnull(dob) else ""
    )
    grouped_df["License Expiration Date"] = grouped_df["License Expiration Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )
    grouped_df["Hire Date"] = grouped_df["Hire Date"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )

    # Add columns that do not have mappings and set them to None
    grouped_df["Years of Experience"] = None
    # grouped_df["Hire Date"] = None # we will pick from drivers list
    grouped_df["Years of Tenure"] = None
    grouped_df["MVR Received"] = "FALSE"
    # grouped_df["undesirable"] = None
    grouped_df["Excluded"] = None
    # grouped_df["prohibited"] = None
    # grouped_df["mvr_score"] = None
    grouped_df["Status"] = None
    # grouped_df["Violation Description"] = None
    grouped_df["Number of Points"] = 0
    grouped_df["FullTime"] = 1
    grouped_df["Comments"] = None
    grouped_df['driver_list_flag'] = None
    grouped_df['mvr_list_flag'] = None
    # Reorder columns to match the required order (using label keys)
    column_order = [
        "Date MVR Ordered",
        "Driver Full Name",
        "Driver First Name",
        "Driver Last Name",
        "Driver Date of Birth",
        "Age",
        "Hire Date",
        "Years of Tenure",
        "Years of Experience",
        "MVR Received",
        "License Number",
        "License Type",
        "License State",
        "License Expiration Date",
        "License Status",
        "Status",
        "Violation Description",
        "Violation Category",
        "Accident Count",
        "Minor Count",
        "Major Count",
        "Minor Count 3 Year",
        "Minor Violation Description 3 Year",
        "Major Count 1 Year",  # only used for post logics will be dropped later
        "Major Violation Description 1 Year",
        "Major Violation Description 5 Year",  # only used for post logics will be dropped later
        "Total Incidents",
        "Excluded",
        "Number of Points",
        "FullTime",
        "Medical Expiration Date",
        "Medical Status",
        "Restrictions",
        "Comments",
        "driver_list_flag",  # this column is just for supporting post logic
        "mvr_list_flag",  # this column is just for supporting post logic
    ]
    #adddelete date
        # Get proposed_effective_date from original data
    proposed_effective_date = ""
    if not df.empty and 'proposed_effective_date' in df.columns:
        valid_dates = df['proposed_effective_date'].dropna()
        if not valid_dates.empty:
            proposed_effective_date = pd.to_datetime(valid_dates.iloc[0]).strftime("%m/%d/%Y")
    
    # Reorder columns first
    grouped_df = grouped_df[column_order]
    
    # Add AddDeleteDate column after Date MVR Ordered
    grouped_df["Proposed Effective Date"] = proposed_effective_date
    
    # Calculate initial Status and Comments based on MVR data
    if vehicle_df is not None and not vehicle_df.empty and "Vehicle Body Type" in vehicle_df.columns and not vehicle_df["Vehicle Body Type"].isnull().all():
        has_heavy_vehicle = vehicle_df["Vehicle Body Type"].str.contains(r"heavy|extra heavy", case=False,
                                                                         na=False).any()
    else:
        has_heavy_vehicle = False

    grouped_df[["Status", "Comments"]] = grouped_df.apply(lambda row: status_and_comments(row, has_heavy_vehicle),
                                                          axis=1)

    grouped_df["Date MVR Ordered"] = grouped_df["Date MVR Ordered"].apply(
        lambda expd: expd.strftime("%m/%d/%Y") if pd.notnull(expd) else ""
    )



    # Match with Driver List
    if driver_df is not None and not driver_df.empty:
        # Ensure first and last name columns exist
        if not (driver_df.replace('', pd.NA).isna().all().all()):
            if 'Driver Full Name' in driver_df.columns:
                driver_df[["Driver First Name", "Driver Last Name"]] = driver_df["Driver Full Name"].apply(
                    split_driver_name)
            
            # Setup columns for driver_df
            driver_df["MVR Received"] = "FALSE"
            driver_df["Number of Points"] = 0
            driver_df["FullTime"] = 1
            driver_df["Status"] = "Pending"
            driver_df["Years of Tenure"] = driver_df["Hire Date"].apply(calculate_age)
            
            # Ensure column presence
            driver_df = driver_df.reindex(columns=column_order, fill_value="")
            
            # Get superset (formatted driver list)
            merged_drivers_df = superset_drivers(driver_df)
            
            # Recalculate Age for merged dataframe
            merged_drivers_df["Age"] = merged_drivers_df["Driver Date of Birth"].apply(calculate_age)

            ''' match MVR data with Driver List '''
            matched_grouped_indices = set()
            
            for i, driver_row in merged_drivers_df.iterrows():
                # Only consider unmatched grouped_df rows (MVR rows)
                unmatched_grouped_df = grouped_df.loc[~grouped_df.index.isin(matched_grouped_indices)]
                
                # Step 1: Match by license number
                license_matches = unmatched_grouped_df.apply(
                    lambda grouped_row: match_by_license(driver_row, grouped_row), axis=1
                )
                matching_row = unmatched_grouped_df[license_matches]
                
                # Step 2: Fallback to name+dob
                if matching_row.empty:
                    fallback_matches = unmatched_grouped_df.apply(
                        lambda grouped_row: match_by_name_dob(driver_row, grouped_row), axis=1
                    )
                    matching_row = unmatched_grouped_df[fallback_matches]
                
                if not matching_row.empty:
                    # Found a match: MVR exists for this driver
                    matched_grouped_indices.add(matching_row.index[0])
                    match_row_data = matching_row.iloc[0]
                    
                    for col in merged_drivers_df.columns:
                        # Skip flags that we set in superset_drivers
                        if col in ["driver_list_flag", "mvr_list_flag"]:
                            continue
                            
                        if col == "mvr_list_flag": # Handled skip above but just structurally logic check
                            pass
                        
                        elif col == "MVR Received":
                            merged_drivers_df.loc[i, col] = "TRUE"
                            
                        elif col == "Years of Tenure":
                            # Calculate logic if needed, or preserve
                            pass
                        elif col == "mvr_list_flag": # Should set to True since we found a match
                             pass # We will set it explicitly below if we want, but logic later uses matching
                        
                        else:
                            # Copy data from MVR (grouped_df) to Driver Row
                            # Except maybe we want to keep Driver List name/DOB?
                            # Usually MVR data overwrites or fills application data?
                            # Existing logic overwrote.
                            merged_drivers_df.loc[i, col] = match_row_data[col]
                            
                    merged_drivers_df.loc[i, "mvr_list_flag"] = True
                else:
                    # No MVR match
                    pass

                        # Add Proposed Effective Date to matched drivers
            merged_drivers_df["Proposed Effective Date"] = proposed_effective_date
            mvr_data = merged_drivers_df.to_dict(orient="records")
            
            # Add unmatched MVR rows (MVRs that didn't match any driver in the list)
            unmatched_grouped_df = grouped_df.loc[~grouped_df.index.isin(matched_grouped_indices)]
            if not unmatched_grouped_df.empty:
                # Set flags for unmatched MVRs
                unmatched_rows = unmatched_grouped_df.copy()
                unmatched_rows['driver_list_flag'] = False
                unmatched_rows['mvr_list_flag'] = True
                                # Add Proposed Effective Date to unmatched MVRs
                unmatched_rows["Proposed Effective Date"] = proposed_effective_date
                mvr_data.extend(unmatched_rows.to_dict(orient="records"))

            ''' Post logic '''
            mvr_data = postprocess_mvr_data(mvr_data)

            # Cleanup
            mvr_data = filter_and_cleanup_mvr_data(mvr_data)
        else:
            # If Driver Df is empty/invalid, just return MVR data
                        # Add Proposed Effective Date when no driver list
            grouped_df["Proposed Effective Date"] = proposed_effective_date
            mvr_data = grouped_df.to_dict(orient="records")
    else:
        # If no driver list provided
        mvr_data = grouped_df.to_dict(orient="records")
        
    return mvr_data


def generate_report(output_dict, driver_df=None, vehicle_df=None, workbook=None, xlsx_sheet_tabs=[]):
    if not workbook:
        workbook = Workbook()

    result_df = pd.DataFrame(output_dict)

    mvr_output = generate_mvr_data_sheet_for_drivers(result_df, driver_df, vehicle_df)
    
    if mvr_output:
        mvr_output_df = pd.DataFrame(mvr_output)
        mvr_output_final = mvr_output_df.to_dict(orient="records")

        write_excel_sheet_mvr_owned(
            mvr_output_final, workbook, "Riscom MVR", "Riscom MVR", False, start_row=2, skip_blue_header=True
        )

    # reorder sheets
    desired_order = ["Riscom MVR", "drivers", "vehicles", "MVR raw"]
    # Only include sheets that actually exist
    ordered_sheets = [name for name in desired_order if name in workbook.sheetnames]
    # Add any other sheets that might exist but are not in the desired order
    ordered_sheets += [name for name in workbook.sheetnames if name not in ordered_sheets]
    workbook._sheets = [workbook[name] for name in ordered_sheets]
    
   
    worksheet = workbook['MVR raw']
    workbook.active = worksheet
    report_buffer = BytesIO()
    workbook.save(report_buffer)
    report_buffer.seek(0)

    return report_buffer, mvr_output_final

def process_riscom_mvr_data(mvr_file_buffer, original_wb):
    df = pd.read_excel(mvr_file_buffer, sheet_name="MVR raw", skiprows=1, dtype={'license_number': str})
    # Reset buffer position for subsequent reads if necessary, though read_excel usually handles it. 
    # But read_excel might not consume it all if sheet_name is specified? It reads the whole file usually.
    # Safe to seek 0.
    mvr_file_buffer.seek(0)
    driver_df = pd.read_excel(mvr_file_buffer, sheet_name="drivers", skiprows=1, dtype={'License Number': str})
    
    vehicle_df = None

    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    sheets_to_keep = ["MVR raw", "drivers"]
    sheets_to_remove = [sheet for sheet in original_wb.sheetnames if sheet not in sheets_to_keep]
    
    for sheet_name in sheets_to_remove:
        if sheet_name in original_wb:
            original_wb.remove(original_wb[sheet_name])
            
    output_dict = df.to_dict(orient="records")
    
    buffer, processed_data = generate_report(output_dict, driver_df, vehicle_df, workbook=original_wb)
    return buffer, processed_data

def download_report_mvr_renewal_riscom_test2(uploaded_file):
    # Check if zip
    mvr_file_buffer = None
    original_wb = None
    
    # Check for zip signature
    is_zip = False
    try:
        if zipfile.is_zipfile(uploaded_file):
            is_zip = True
    except Exception:
        pass # Might be BytesIO acting up or empty

    if is_zip:
        zip_buffer = io.BytesIO(uploaded_file.read())
        with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
            for file_name in zip_file.namelist():
                if file_name.startswith('__MACOSX/') or file_name.startswith('._'):
                    continue
                base_name = os.path.basename(file_name.lower())
                if base_name.startswith('mvr'):
                    with zip_file.open(file_name) as file:
                        file_content = file.read()
                    original_wb = load_workbook(io.BytesIO(file_content))
                    mvr_file_buffer = io.BytesIO(file_content)
                    break
    else:
        # Assume Excel
        try:
            uploaded_file.seek(0)
            file_content = uploaded_file.read()
            mvr_file_buffer = io.BytesIO(file_content)
            original_wb = load_workbook(mvr_file_buffer)
            mvr_file_buffer.seek(0)
        except Exception:
             # If it fails, maybe it's not a valid Excel
             pass

    if mvr_file_buffer is None or original_wb is None:
        raise Exception("MVR file not found or invalid format. Please upload a valid Excel or Zip file containing 'report_mvr...'.")

    buffer, _ = process_riscom_mvr_data(mvr_file_buffer, original_wb)
    return [buffer]


if __name__ == "__main__":
    file_path = '/home/dell/Desktop/pibit/insight-board/Archive (26).zip'

    try:
        with open(file_path, "rb") as uploaded_file:
            print(f"Processing file: {file_path}")

            # Keep original uploaded name
            uploaded_filename = os.path.basename(uploaded_file.name)
            uploaded_name_no_ext, _ = os.path.splitext(uploaded_filename)

            result = download_report_mvr_renewal_riscom_test2(uploaded_file)

            # Case 1: ZIP file → extract actual MVR filename
            if zipfile.is_zipfile(file_path):
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    for file_name in zip_file.namelist():
                        if file_name.startswith('__MACOSX/') or file_name.startswith('._'):
                            continue

                        base_name = os.path.basename(file_name)
                        base_name_lower = base_name.lower()

                        if base_name_lower.startswith('mvr'):
                            file_name_without_ext, _ = os.path.splitext(base_name)
                            break
                    else:
                        file_name_without_ext = uploaded_name_no_ext
            else:
                # Direct Excel upload
                file_name_without_ext = uploaded_name_no_ext

            # Remove "report_" prefix if exists
            if file_name_without_ext.lower().startswith("report_"):
                file_name_without_ext = file_name_without_ext[len("report_"):]

            output_filename = f"{file_name_without_ext}.xlsx"

            with open(output_filename, "wb") as f:
                f.write(result[0].getvalue())

            print(f"Successfully generated report: {output_filename}")

    except FileNotFoundError:
        print(f"Error: Test file not found at {file_path}")
    except Exception as e:
        print(f"An error occurred during processing: {e}")
        import traceback
        traceback.print_exc()
