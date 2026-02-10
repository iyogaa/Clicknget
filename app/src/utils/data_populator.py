import openpyxl
from openpyxl.utils import column_index_from_string
import logging
from typing import Optional, Any
import os

# Set up logging
logger = logging.getLogger(__name__)

from datetime import datetime

def _get_value_nearby(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> Optional[Any]:
    """
    Extract the value from the adjacent column (right side) OR the cell immediately below.
    """
    # Try right
    val_right = sheet.cell(row=row, column=col + 1).value
    if val_right is not None and str(val_right).strip():
        return val_right
    
    # Try below
    val_below = sheet.cell(row=row + 1, column=col).value
    if val_below is not None and str(val_below).strip():
        return val_below
    
    return None

def _clean_extracted_value(val: Any) -> Any:
    """
    Clean extracted values. If it's a string, trim and remove colons.
    If it's a datetime, return as is.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        s = val.strip()
        s = s.strip(':').strip()
        return s
    return val

def populate_client_data(client_file_path: str, report_file_path: str, na_value: str = "") -> None:
    """
    Reads client Excel files and populates specific cells in a final report Excel file.
    """
    if not os.path.exists(client_file_path):
        logger.warning(f"Client file not found: {client_file_path}")
        return

    if not os.path.exists(report_file_path):
        logger.warning(f"Report file not found: {report_file_path}")
        return

    try:
        # Load client workbook (read-only)
        client_wb = openpyxl.load_workbook(client_file_path, data_only=True, read_only=True)
        sheet = client_wb.active
        
        insured_name = None
        effective_dates = None
        
        # Labels to search for
        insured_labels = ["insured name", "entity name"]
        # Handle variations: "Effective Date", "Effective Dates"
        effective_labels = ["effective date", "effective dates"]
        
        def find_in_row(row_idx, start_col=1, end_col=None):
            nonlocal insured_name, effective_dates
            if not end_col: end_col = min(sheet.max_column + 1, 50)
            
            for c_idx in range(start_col, end_col):
                cell_val = sheet.cell(row=row_idx, column=c_idx).value
                if cell_val is None: continue
                
                s_val = str(cell_val).lower().strip()
                
                if not insured_name:
                    if any(label in s_val for label in insured_labels):
                        # Prioritize "Insured name"
                        if "insured name" in s_val or insured_name is None:
                            val = _get_value_nearby(sheet, row_idx, c_idx)
                            if val: insured_name = _clean_extracted_value(val)
                
                if not effective_dates:
                    if any(label in s_val for label in effective_labels):
                        val = _get_value_nearby(sheet, row_idx, c_idx)
                        if val: effective_dates = _clean_extracted_value(val)

        # 1. Search Column A first
        for r in range(1, min(sheet.max_row + 1, 100)):
            find_in_row(r, start_col=1, end_col=2)
            if insured_name and effective_dates: break

        # 2. Scan entire sheet if not found
        if not insured_name or not effective_dates:
            for r in range(1, min(sheet.max_row + 1, 100)):
                find_in_row(r, start_col=2) # Skip col A
                if insured_name and effective_dates: break
                            
        client_wb.close()
        
        # Fallback handling
        final_insured_name = insured_name if insured_name is not None else na_value
        final_effective_dates = effective_dates if effective_dates is not None else na_value
            
        # Load report workbook to modify
        report_wb = openpyxl.load_workbook(report_file_path)
        report_sheet = report_wb.active
        
        # Populate Target Cells
        report_sheet['C1'] = final_insured_name
        report_sheet['C2'] = final_effective_dates
        
        # Save changes
        report_wb.save(report_file_path)
        logger.info(f"Successfully populated client data from {client_file_path} into {report_file_path}")
        
    except PermissionError:
        logger.error(f"Permission denied when accessing files.")
    except Exception as e:
        logger.error(f"Error during client data population: {e}", exc_info=True)
        
        # Log if not found
        if not final_insured_name:
            logger.warning(f"Could not find Insured Name in {client_file_path}")
        if not final_effective_dates:
            logger.warning(f"Could not find Effective Dates in {client_file_path}")
            
        # Load report workbook to modify
        report_wb = openpyxl.load_workbook(report_file_path)
        report_sheet = report_wb.active # Assuming the first/active sheet is intended
        
        # Populate Target Cells
        report_sheet['C1'] = final_insured_name
        report_sheet['C2'] = final_effective_dates
        
        # Save changes
        report_wb.save(report_file_path)
        logger.info(f"Successfully populated client data from {client_file_path} into {report_file_path}")
        
    except PermissionError:
        logger.error(f"Permission denied when accessing files.")
    except Exception as e:
        logger.error(f"Error during client data population: {e}", exc_info=True)
