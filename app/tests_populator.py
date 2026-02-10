import pytest
import openpyxl
import os
from app.src.utils.data_populator import populate_client_data

def test_extraction_and_population(tmp_path):
    # 1. Create a mock client file
    client_path = str(tmp_path / "client.xlsx")
    c_wb = openpyxl.Workbook()
    c_ws = c_wb.active
    c_ws['A5'] = "Insured name:"
    c_ws['B5'] = "John Doe Enterprises"
    c_ws['A10'] = "Effective Dates:"
    c_ws['A11'] = "01/01/2024 - 01/01/2025"
    c_wb.save(client_path)
    
    # 2. Create a mock report file
    report_path = str(tmp_path / "report.xlsx")
    r_wb = openpyxl.Workbook()
    r_ws = r_wb.active
    r_wb.save(report_path)
    
    # 3. Call the populator
    populate_client_data(client_path, report_path)
    
    # 4. Verify results
    v_wb = openpyxl.load_workbook(report_path)
    v_ws = v_wb.active
    assert v_ws['C1'].value == "John Doe Enterprises"
    assert v_ws['C2'].value == "01/01/2024 - 01/01/2025"
    v_wb.close()

def test_extraction_variations(tmp_path):
    client_path = str(tmp_path / "client_variations.xlsx")
    c_wb = openpyxl.Workbook()
    c_ws = c_wb.active
    c_ws['C3'] = "ENTITY NAME:" # Partial/Case variation
    c_ws['C4'] = "Acme Corp" # Value below
    c_ws['E10'] = "Effective Date"
    c_ws['F10'] = "2024-05-20" # Value right
    c_wb.save(client_path)
    
    report_path = str(tmp_path / "report_variations.xlsx")
    r_wb = openpyxl.Workbook()
    r_wb.save(report_path)
    
    populate_client_data(client_path, report_path)
    
    v_wb = openpyxl.load_workbook(report_path)
    v_ws = v_wb.active
    assert v_ws['C1'].value == "Acme Corp"
    assert v_ws['C2'].value == "2024-05-20"
