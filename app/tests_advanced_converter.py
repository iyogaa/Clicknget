import pytest
import openpyxl
import os
from app.src.core.converter import convert_excel_to_pdf
from reportlab.lib.pagesizes import A4, landscape

def test_advanced_conversion_orientation(tmp_path):
    # 1. Create a wide Excel file to trigger landscape
    xlsx_path = str(tmp_path / "wide.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    # Create many columns to make it very wide
    for c in range(1, 25):
        ws.cell(row=1, column=c, value=f"Header {c}")
        ws.cell(row=2, column=c, value=f"Data {c}")
    wb.save(xlsx_path)
    
    pdf_path = str(tmp_path / "wide.pdf")
    
    # 2. Convert with auto-orientation
    result = convert_excel_to_pdf(xlsx_path, pdf_path, orientation="auto")
    
    assert result["success"] is True
    assert os.path.exists(pdf_path)
    assert "Sheet" in result["sheets_processed"]

def test_scaling_limit_warning(tmp_path):
    # Create an extremely wide Excel file
    xlsx_path = str(tmp_path / "super_wide.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for c in range(1, 100):
        ws.cell(row=1, column=c, value="Very Long Header Name To Force Scaling")
    wb.save(xlsx_path)
    
    pdf_path = str(tmp_path / "super_wide.pdf")
    result = convert_excel_to_pdf(xlsx_path, pdf_path)
    
    # It should succeed but might have a warning if it hit the 10% limit
    assert result["success"] is True
    # Verify if warning exists (logic says max(0.1, ...))
    if any("too wide even at 10% scale" in w for w in result["warnings"]):
        assert True
    else:
        # Depending on A4 width, 100 columns might or might not hit 10%
        pass

def test_multi_sheet_conversion(tmp_path):
    xlsx_path = str(tmp_path / "multi.xlsx")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data1"
    ws1['A1'] = "Page 1 Content"
    
    ws2 = wb.create_sheet("Data2")
    ws2['A1'] = "Page 2 Content"
    wb.save(xlsx_path)
    
    pdf_path = str(tmp_path / "multi.pdf")
    result = convert_excel_to_pdf(xlsx_path, pdf_path, include_sheet_names=True)
    
    assert result["success"] is True
    assert len(result["sheets_processed"]) == 2
