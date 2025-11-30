import io
import os
from copy import copy
from datetime import date
import logging
import re
import argparse
import sys

import pandas as pd
from openpyxl import load_workbook
import difflib

class Alltrans:
    def __init__(self, template_path="Template.xlsx",
                 alltrans_sheet="All Trans", alltrans_header_row=4, mvr_sheet_name="MVR"):
        self.TEMPLATE_PATH = template_path
        self.ALLTRANS_SHEET = alltrans_sheet
        self.ALLTRANS_HEADER_ROW = alltrans_header_row
        self.MVR_PREFERRED_NAME = mvr_sheet_name
        # logger for debugging instead of silent excepts
        self.logger = logging.getLogger(self.__class__.__name__)
        if not self.logger.handlers:
            # don't configure root logger here; add a simple StreamHandler if none
            handler = logging.StreamHandler()
            handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(name)s: %(message)s'))
            self.logger.addHandler(handler)
        self.logger.setLevel(logging.INFO)

    # ---------------- Helpers ----------------
    def _replicate_sheet_across_workbooks(self, src_ws, dst_wb, dst_title):
        if dst_title in dst_wb.sheetnames:
            dst_wb.remove(dst_wb[dst_title])
        dst_ws = dst_wb.create_sheet(dst_title)
        for row in src_ws.iter_rows(values_only=False):
            for src_cell in row:
                r = src_cell.row
                c = src_cell.column
                dst_cell = dst_ws.cell(row=r, column=c, value=src_cell.value)
                try:
                    if src_cell.has_style:
                        if src_cell.font is not None:
                            dst_cell.font = copy(src_cell.font)
                        if src_cell.fill is not None:
                            dst_cell.fill = copy(src_cell.fill)
                        if src_cell.border is not None:
                            dst_cell.border = copy(src_cell.border)
                        if src_cell.alignment is not None:
                            dst_cell.alignment = copy(src_cell.alignment)
                        if src_cell.number_format is not None:
                            dst_cell.number_format = src_cell.number_format
                except Exception as e:
                    self.logger.debug("Failed copying style from cell %s:%s — %s", r, c, e, exc_info=True)
                try:
                    if src_cell.hyperlink:
                        dst_cell._hyperlink = copy(src_cell.hyperlink)
                except Exception as e:
                    self.logger.debug("Failed copying hyperlink for cell %s:%s — %s", r, c, e, exc_info=True)
                try:
                    if src_cell.comment:
                        dst_cell.comment = copy(src_cell.comment)
                except Exception as e:
                    self.logger.debug("Failed copying comment for cell %s:%s — %s", r, c, e, exc_info=True)
        try:
            for merged in list(src_ws.merged_cells.ranges):
                dst_ws.merge_cells(str(merged))
        except Exception as e:
            self.logger.debug("Failed copying merged cells: %s", e, exc_info=True)
        try:
            for col_letter, col_dim in src_ws.column_dimensions.items():
                if col_dim.width is not None:
                    dst_ws.column_dimensions[col_letter].width = col_dim.width
        except Exception as e:
            self.logger.debug("Failed copying column dimensions: %s", e, exc_info=True)
        try:
            for r_idx, row_dim in src_ws.row_dimensions.items():
                if row_dim.height is not None:
                    dst_ws.row_dimensions[r_idx].height = row_dim.height
        except Exception as e:
            self.logger.debug("Failed copying row dimensions: %s", e, exc_info=True)
        try:
            dst_ws.sheet_view = copy(src_ws.sheet_view)
        except Exception as e:
            self.logger.debug("Failed copying sheet_view: %s", e, exc_info=True)
        try:
            dst_ws.freeze_panes = src_ws.freeze_panes
        except Exception as e:
            self.logger.debug("Failed copying freeze_panes: %s", e, exc_info=True)
        try:
            dst_ws.page_setup = copy(src_ws.page_setup)
            dst_ws.page_margins = copy(src_ws.page_margins)
            dst_ws.print_options = copy(src_ws.print_options)
        except Exception as e:
            self.logger.debug("Failed copying page setup/margins/print options: %s", e, exc_info=True)
        try:
            dst_ws.sheet_properties = copy(src_ws.sheet_properties)
        except Exception as e:
            self.logger.debug("Failed copying sheet_properties: %s", e, exc_info=True)
        try:
            dst_ws.protection = copy(src_ws.protection)
        except Exception as e:
            self.logger.debug("Failed copying protection: %s", e, exc_info=True)
        try:
            if hasattr(src_ws, "tab_color"):
                dst_ws.tab_color = copy(src_ws.tab_color)
        except Exception as e:
            self.logger.debug("Failed copying tab_color: %s", e, exc_info=True)
        return dst_ws

    def _format_doh_for_excel(self, val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            ts = pd.to_datetime(val, errors="coerce")
            if pd.isna(ts):
                s = str(val).strip()
                return f"'{s}" if s != "" else None
            return f"{ts.strftime('%m/%d/%Y')}"
        except Exception:
            s = str(val).strip()
            return f"'{s}" if s != "" else None


    def _normalize_tokens(self, name):
        if name is None:
            return []
        s = str(name).strip().lower()
        if s == "":
            return []
        for ch in [",", ".", "-", "_", "/"]:
            s = s.replace(ch, " ")
        return [t for t in s.split() if t]

    def _name_token_overlap(self, m_tokens, l_tokens):
        if not m_tokens or not l_tokens:
            return 0
        m_set = set(m_tokens)
        l_set = set(l_tokens)
        exact = len(m_set & l_set)
        initials = 0
        for mt in m_set:
            for lt in l_set:
                if mt and lt and mt[0] == lt[0]:
                    initials += 1
        contain = 0
        for mt in m_set:
            for lt in l_set:
                if mt in lt or lt in mt:
                    contain += 1
        return max(exact, initials, contain)

    def _name_dob_flexible_match(self, mvr_name, mvr_dob, lookup_name, lookup_dob):
        try:
            m_dt = pd.to_datetime(mvr_dob, errors="coerce")
            l_dt = pd.to_datetime(lookup_dob, errors="coerce")
            if pd.isna(m_dt) or pd.isna(l_dt):
                return False
            if m_dt.date() != l_dt.date():
                return False
        except Exception:
            return False
        m_tokens = self._normalize_tokens(mvr_name)
        l_tokens = self._normalize_tokens(lookup_name)
        if not m_tokens or not l_tokens:
            return False
        overlap = self._name_token_overlap(m_tokens, l_tokens)
        if overlap >= 2:
            return True
        if len(m_tokens) >= 2:
            return (m_tokens[0] in " ".join(l_tokens)) and (m_tokens[-1] in " ".join(l_tokens))
        return m_tokens[0] in " ".join(l_tokens)

    def _score_name_match(self, a, b):
        a_tokens = self._normalize_tokens(a)
        b_tokens = self._normalize_tokens(b)
        if not a_tokens or not b_tokens:
            return 0
        overlap = self._name_token_overlap(a_tokens, b_tokens)
        score = overlap
        a_norm = " ".join(a_tokens)
        b_norm = " ".join(b_tokens)
        if a_norm == b_norm:
            score += 3
        if len(a_tokens) >= 2:
            if (a_tokens[0] in " ".join(b_tokens)) and (a_tokens[-1] in " ".join(b_tokens)):
                score += 2
        return score

    def _extract_first_last(self, name_std):
        if name_std is None:
            return "", ""
        toks = [t for t in str(name_std).split() if t]
        if not toks:
            return "", ""
        first = toks[0]
        i = len(toks) - 1
        last = toks[i]
        connectors = {"DE","LA","DEL","DELA","DA","VAN","VON","DI","AL","BIN","BINTI","MC","MAC","ST"}
        while i - 1 >= 0 and toks[i-1] in connectors:
            last = toks[i-1] + " " + last
            i -= 1
        return first, last

    def _first_last_equal(self, a_std, b_std):
        af, al = self._extract_first_last(a_std)
        bf, bl = self._extract_first_last(b_std)
        return (af == bf and al == bl) or (af == bl and al == bf)

    def _tokens_clean(self, name_std):
        if not name_std:
            return []
        toks = [t for t in str(name_std).split() if t]
        connectors = {"DE","LA","DEL","DELA","DA","VAN","VON","DI","AL","BIN","BINTI","MC","MAC","ST"}
        return [t for t in toks if t not in connectors]

    def _is_name_match(self, name_a, name_b):
        tokens_a = self._tokens_clean(name_a)
        tokens_b = self._tokens_clean(name_b)
        
        if not tokens_a or not tokens_b:
            return False
            
        # Helper for fuzzy token matching
        def get_fuzzy_matches(list_sub, list_super):
            matches = 0
            used_super_indices = set()
            for t_sub in list_sub:
                for i, t_super in enumerate(list_super):
                    if i in used_super_indices:
                        continue
                    
                    # 1. Exact Match
                    if t_sub == t_super:
                        matches += 1
                        used_super_indices.add(i)
                        break
                    
                    # 2. Initial Match (if one is single letter)
                    # e.g. "K" matches "KRISHNAN"
                    if len(t_sub) == 1 and len(t_super) > 1 and t_super.startswith(t_sub):
                         matches += 1
                         used_super_indices.add(i)
                         break
                    # Note: The reverse (t_super is single letter) is handled by the caller swapping lists

                    # 3. Fuzzy Match
                    if difflib.SequenceMatcher(None, t_sub, t_super).ratio() > 0.80:
                        matches += 1
                        used_super_indices.add(i)
                        break
            return matches

        # Check overlap count
        matches_a_in_b = get_fuzzy_matches(tokens_a, tokens_b)
        matches_b_in_a = get_fuzzy_matches(tokens_b, tokens_a)
        
        # Subset condition: All tokens of A found in B, or all tokens of B found in A
        if matches_a_in_b == len(tokens_a) or matches_b_in_a == len(tokens_b):
            return True
            
        # Overlap condition: At least 2 tokens match (e.g. First Name and Last Name)
        if matches_a_in_b >= 2:
            return True
            
        return False

    def _find_cdl_col(self, cols, sample_df=None):
        candidates = ["CDL Number", "CDLNumber", "CDL No", "CDL", "CDL_Number", "License Number", "License Num", "License No", "Driver License Number", "Driver License", "License", "Drivers License Number", "DL Number", "DL No"]
        for c in candidates:
            for col in cols:
                if col.strip().lower() == c.strip().lower():
                    return col
        for col in cols:
            if "cdl" in col.strip().lower() or "license" in col.strip().lower() and "number" in col.strip().lower():
                return col
        if sample_df is not None:
            for col in cols:
                try:
                    sample = sample_df[col].dropna().astype(str)
                    if not sample.empty and sample.iloc[0].strip().replace(" ", "").isdigit():
                        return col
                except Exception:
                    continue
        return None

    def _find_hire_col(self, cols):
        candidates = ["Hire Date", "HireDate", "DOH", "Date of Hire", "Hire_Date", "Driver Hire Date", "Driver Hire date", "Hire date", "Date Hired", "Employment Date", "Start Date", "Hire Date", "Date of Join", "Joining Date", "Join Date"]
        for c in candidates:
            for col in cols:
                if col.strip().lower() == c.strip().lower():
                    return col
        for col in cols:
            if "hire" in col.strip().lower() or "doh" in col.strip().lower() or "date" in col.strip().lower():
                return col
        # last resort: pick any column that has date-like content
        for col in cols:
            try:
                sample = pd.Series([c for c in cols if not pd.isna(c)]).head(1)
                # just return first col as fallback if it exists
                return col
            except Exception:
                continue
        return None

    def _compute_age_from_str_dob(self, s):
        if s is None:
            return None
        try:
            s2 = s[1:] if isinstance(s, str) and s.startswith("'") else s
            ts = pd.to_datetime(s2, errors="coerce")
            if pd.isna(ts):
                return None
            today = date.today()
            return today.year - ts.year - ((today.month, today.day) < (ts.month, ts.day))
        except Exception:
            return None

    def _std_upper(self, s):
        if s is None:
            return ""
        try:
            import unicodedata
            x = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii')
        except Exception:
            x = str(s)
        x = x.upper().strip()
        x = re.sub(r"[_]+", " ", x)
        # Replace non-alphanumeric with space to preserve token separation (e.g. Smith-Jones -> SMITH JONES)
        x = re.sub(r"[^A-Z0-9\s]", " ", x)
        # remove common prefixes/suffixes like MR, MRS, MS, DR, JR, SR, II, III, IV
        tokens = [t for t in re.split(r"\s+", x) if t]
        drop = {"MR","MRS","MS","DR","JR","SR","II","III","IV"}
        tokens = [t for t in tokens if t not in drop]
        x = " ".join(tokens)
        x = re.sub(r"\s+", " ", x).strip()
        return x

    def _dob_iso(self, val):
        if val is None:
            return None
        try:
            v = val
            if isinstance(v, str) and v.startswith("'"):
                v = v[1:]
            ts = pd.to_datetime(v, errors="coerce")
            if pd.isna(ts):
                ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
            if pd.isna(ts):
                return None
            return ts.strftime("%Y-%m-%d")
        except Exception:
            return None

    def _build_full_name(self, row, cols):
        vals = []
        for col in cols:
            if col and col in row and row.get(col) is not None and str(row.get(col)).strip() != "":
                vals.append(self._std_upper(row.get(col)))
        return " ".join([v for v in vals if v]).strip()

    def _date_equal(self, a, b):
        ai = self._dob_iso(a)
        bi = self._dob_iso(b)
        return ai is not None and bi is not None and ai == bi

    def _clean_cdl_key(self, key: str) -> str:
        """Normalize a CDL key by removing unwanted characters, stripping leading zeros
        from the numeric portion, and standardizing a two-letter state code to the
        suffix form 'NUMBER-SS' (e.g. '012345-CA'). Returns empty string when
        nothing usable remains.
        """
        if key is None:
            return ""
        s = str(key).strip().upper()
        if s == "":
            return ""
        # remove all non-alphanumeric characters
        s2 = re.sub(r'[^A-Z0-9]', '', s)
        if s2 == "":
            return ""
        # heuristic: correct common OCR confusions before digit extraction
        ocr_map = str.maketrans({
            'O': '0', 'I': '1', 'L': '1', 'S': '5', 'B': '8', 'Z': '2', 'G': '6'
        })
        s2 = s2.translate(ocr_map)
        # detect two-letter state at start or end
        state = ""
        num = s2
        if len(s2) >= 2 and s2[:2].isalpha():
            state = s2[:2]
            num = s2[2:]
        elif len(s2) >= 2 and s2[-2:].isalpha():
            state = s2[-2:]
            num = s2[:-2]
        else:
            m = re.search(r'([A-Z]{2})', s2)
            if m:
                state = m.group(1)
                num = s2.replace(state, '')
        # keep only digits from numeric part
        num_digits = ''.join(re.findall(r'\d+', num))
        num_digits = num_digits.lstrip('0')
        if num_digits == "":
            # if no digits remain, fall back to the raw letters (if any)
            if state:
                return state
            return ""
        if state:
            return f"{num_digits}-{state}"
        return num_digits

    # ---------------- Core Logic ----------------
    def process_data(self, main_bytes: bytes, lookup_bytes: bytes, chosen_lookup_sheet: str = None, preview_rows: int = 8):
        # load main workbook (openpyxl) for replication
        main_wb = load_workbook(io.BytesIO(main_bytes), data_only=False)
        if self.MVR_PREFERRED_NAME in main_wb.sheetnames:
            src_mvr_ws = main_wb[self.MVR_PREFERRED_NAME]
            mvr_sheet_name_used = self.MVR_PREFERRED_NAME
        else:
            src_mvr_ws = main_wb[main_wb.sheetnames[0]]
            mvr_sheet_name_used = main_wb.sheetnames[0]

        # read raw main MVR to apply skip rules
        df_raw = pd.read_excel(io.BytesIO(main_bytes), sheet_name=mvr_sheet_name_used, dtype=str, header=None)
        if len(df_raw) < 2:
            raise ValueError("Main MVR sheet too small to process.")
        # auto-detect header row in the first few rows
        def _score_mvr_header(row):
            text_vals = [str(x).strip().lower() for x in row.fillna("").astype(str).tolist()]
            score = 0
            tokens = ("cdl", "driver", "name", "dob", "hire", "date")
            for t in text_vals:
                for ex in tokens:
                    if ex in t:
                        score += 1
            return score

        scan_rows = min(6, len(df_raw))
        header_scores = [_score_mvr_header(df_raw.iloc[i]) for i in range(scan_rows)]
        best_idx = int(pd.Series(header_scores).idxmax())
        best_score = header_scores[best_idx]
        header_row_idx = best_idx if best_score >= 1 else 1
        header_row_mvr = df_raw.iloc[header_row_idx].fillna("").astype(str).tolist()
        data_rows = df_raw.drop(index=range(0, header_row_idx + 1)).reset_index(drop=True)
        # drop rows that are entirely empty
        data_rows = data_rows.dropna(how="all").reset_index(drop=True)
        df_main_clean = data_rows.copy()
        df_main_clean.columns = [h.strip() if h is not None else "" for h in header_row_mvr]
        df_main_clean.columns = [str(c).strip() for c in df_main_clean.columns]

        # lookup workbook and sheet selection
        lookup_wb = load_workbook(io.BytesIO(lookup_bytes), read_only=True, data_only=True)
        lookup_sheets = lookup_wb.sheetnames
        if chosen_lookup_sheet and chosen_lookup_sheet in lookup_sheets:
            chosen_sheet = chosen_lookup_sheet
        else:
            chosen_sheet = lookup_sheets[0]

        # preview and auto-detect header row
        df_lookup_preview = pd.read_excel(io.BytesIO(lookup_bytes), sheet_name=chosen_sheet, dtype=str, header=None, nrows=preview_rows)
        def score_row_as_header(row):
            text_vals = [str(x).strip().lower() for x in row.fillna("").astype(str).tolist()]
            score = 0
            tokens = ("cdl","hire","doh","driver","name","dob","birth")
            for t in text_vals:
                for ex in tokens:
                    if ex in t:
                        score += 1
            return score
        header_scores = [score_row_as_header(df_lookup_preview.iloc[i]) for i in range(len(df_lookup_preview))]
        best_idx = int(pd.Series(header_scores).idxmax())
        best_score = header_scores[best_idx]
        auto_header_row = best_idx if best_score >= 2 else 0
        lookup_df = pd.read_excel(io.BytesIO(lookup_bytes), sheet_name=chosen_sheet, dtype=str, header=auto_header_row)
        lookup_df.columns = [str(c).strip() for c in lookup_df.columns]

        lookup_cols = list(lookup_df.columns)
        cdl_col_lookup = self._find_cdl_col(lookup_cols, sample_df=lookup_df)
        hire_col_lookup = self._find_hire_col(lookup_cols)
        if hire_col_lookup is None:
            raise ValueError("Could not detect Hire Date column in lookup automatically.")
        
        if cdl_col_lookup is None:
            self.logger.warning("Could not detect CDL column in lookup. CDL matching will be skipped.")

        # normalize lookup keys and detect duplicates
        if cdl_col_lookup:
            lookup_df["_cdl_key"] = lookup_df[cdl_col_lookup].astype(str).str.strip()
        else:
            lookup_df["_cdl_key"] = ""

        dup_mask = lookup_df["_cdl_key"].duplicated(keep=False)
        dupe_rows = lookup_df[dup_mask].copy()
        dupe_summary = dupe_rows["_cdl_key"].value_counts().to_dict() if not dupe_rows.empty else {}
        lookup_small = lookup_df.drop_duplicates("_cdl_key", keep="first").copy()
        lookup_rows = lookup_df.to_dict(orient="records")

        for r in lookup_rows:
            r["_cdl_key_norm"] = str(r.get("_cdl_key", "")).strip()
            
            # Identify name columns more carefully
            cols_lower = {c: str(c).lower().strip() for c in lookup_df.columns}
            
            first_col = next((c for c, cl in cols_lower.items() if "first" in cl and "name" in cl), None)
            if not first_col: first_col = next((c for c, cl in cols_lower.items() if "first" in cl), None)
            
            middle_col = next((c for c, cl in cols_lower.items() if "middle" in cl), None)
            
            last_col = next((c for c, cl in cols_lower.items() if "last" in cl and "name" in cl), None)
            if not last_col: last_col = next((c for c, cl in cols_lower.items() if "last" in cl), None)
            
            # Explicit Full Name column candidates
            full_name_col = next((c for c, cl in cols_lower.items() if "full" in cl and "name" in cl), None)
            if not full_name_col:
                full_name_col = next((c for c, cl in cols_lower.items() if "driver" in cl and "name" in cl and "first" not in cl and "last" not in cl), None)
            
            # Generic Name column (fallback)
            generic_name_col = next((c for c, cl in cols_lower.items() if "name" in cl and "first" not in cl and "last" not in cl and "middle" not in cl and "file" not in cl), None)

            parts = []
            if first_col: parts.append(first_col)
            if middle_col: parts.append(middle_col)
            if last_col: parts.append(last_col)
            
            built_name = self._build_full_name(r, parts) if parts else None
            raw_full_name = r.get(full_name_col) if full_name_col else None
            raw_generic_name = r.get(generic_name_col) if generic_name_col else None
            
            # Priority: 
            # 1. Explicit Full Name column
            # 2. Constructed from First/Last
            # 3. Generic "Name" column
            use_name = None
            if raw_full_name and str(raw_full_name).strip():
                use_name = raw_full_name
            elif built_name and str(built_name).strip():
                use_name = built_name
            elif raw_generic_name and str(raw_generic_name).strip():
                use_name = raw_generic_name
            
            r["_lookup_name"] = use_name
            r["_lookup_fullname_std"] = self._std_upper(use_name) if use_name else ""
            
            dob_col = next((c for c in lookup_df.columns if any(k in str(c).lower() for k in ("dob","date of birth","birth"))), None)
            dob_val = r.get(dob_col) if dob_col else None
            try:
                dt = pd.to_datetime(dob_val, errors="coerce")
                r["_lookup_dob"] = dt.strftime("%m/%d/%Y") if not pd.isna(dt) else None
            except Exception:
                r["_lookup_dob"] = None
            r["_lookup_dob_iso"] = self._dob_iso(dob_val)
            r["_hire_raw"] = r.get(hire_col_lookup)

        # build direct CDL map
        lookup_map = pd.Series(lookup_small[hire_col_lookup].values, index=lookup_small["_cdl_key"]).to_dict()
        # build cleaned lookup map and index mapping to support fallback cleaned key matching
        def _clean_key_local(k):
            try:
                return self._clean_cdl_key(str(k))
            except Exception:
                return str(k).strip()

        lookup_map_clean = {}
        lookup_key_to_indices = {}
        for i, lr in enumerate(lookup_rows):
            kraw = str(lr.get("_cdl_key_norm", "")).strip()
            kclean = _clean_key_local(kraw)
            if kclean:
                # prefer first occurrence for map
                lookup_map_clean.setdefault(kclean, lr.get(hire_col_lookup))
            # index mapping for finding matched rows
            lookup_key_to_indices.setdefault(kraw, []).append(i)
            lookup_key_to_indices.setdefault(kclean, []).append(i)

        # prepare main keys & order
        cdl_col_main = self._find_cdl_col(list(df_main_clean.columns), sample_df=df_main_clean)
        
        # Assign unique keys for missing CDL to prevent grouping distinct drivers
        def _make_key(val, idx):
            s = str(val).strip()
            if s == "" or s.lower() == "nan":
                return f"_NO_CDL_{idx}"
            return s

        if cdl_col_main:
            df_main_clean["_cdl_key"] = [
                _make_key(x, i) 
                for i, x in enumerate(df_main_clean[cdl_col_main])
            ]
        else:
            # If no CDL column found, treat all as missing CDL
            self.logger.warning("No CDL column found in MVR. Proceeding with Name+DOB matching.")
            df_main_clean["_cdl_key"] = [f"_NO_CDL_{i}" for i in range(len(df_main_clean))]
        
        order_keys = df_main_clean["_cdl_key"].dropna().astype(str).tolist()
        seen = set()
        ordered_unique_keys = []
        for k in order_keys:
            if k not in seen:
                seen.add(k)
                ordered_unique_keys.append(k)

        # pick main columns
        cols = list(df_main_clean.columns)
        def pick(col_candidates):
            for cand in col_candidates:
                for c in cols:
                    if c.strip().lower() == cand.strip().lower():
                        return c
            for cand in col_candidates:
                for c in cols:
                    if cand.strip().lower() in c.strip().lower():
                        return c
            return None

        driver_col = pick(["Driver Full Name","DriverFullName","Name of Driver","Driver"])
        first_col_main = pick(["First Name","First","Driver First Name"]) or next((c for c in cols if "first" in str(c).lower()), None)
        middle_col_main = pick(["Middle Name","Middle"]) or next((c for c in cols if "middle" in str(c).lower()), None)
        last_col_main = pick(["Last Name","Last","Driver Last Name"]) or next((c for c in cols if "last" in str(c).lower()), None)
        dob_col = pick(["Driver Date of Birth","Date of birth","DOB"])
        cdl_type_col = pick(["CDL Type","CDLType","Lic Class","License Class"])
        lic_state_col = pick(["License State","LicenseState","State"])
        pass_end_col = pick(["License Endorsement - Passenger","Passenger Endorsement","Passenger Endt"])
        school_end_col = pick(["License Endorsement - School Bus","School Bus Endorsement","School Bus Endt"])
        viol_desc_col = pick(["Violation Description","ViolationDescription","Violation Desc","Description"])
        viol_cat_col = pick(["Violation Category","Violation Type","Category"])
        lic_status_col = pick(["License Status","Lic Status","Status"])

        # aggregate per CDL and normalize MVR DOB to MM/DD/YYYY with leading apostrophe
        grouped = df_main_clean.groupby("_cdl_key", sort=False)
        records = []
        for key in ordered_unique_keys:
            group = grouped.get_group(key) if key in grouped.groups else pd.DataFrame(columns=df_main_clean.columns)
            rec = {}
            rec["_cdl_key"] = key
            name_val = group[driver_col].dropna().astype(str).iloc[0] if (driver_col in group and not group[driver_col].dropna().empty) else None
            if not name_val:
                row0 = group.iloc[0].to_dict() if len(group) > 0 else {}
                name_val = self._build_full_name(row0, [first_col_main, middle_col_main, last_col_main]) if any([first_col_main, middle_col_main, last_col_main]) else None
            rec["Driver Full Name"] = name_val
            raw_dob = group[dob_col].dropna().iloc[0] if (dob_col in group and not group[dob_col].dropna().empty) else None
            try:
                dt = pd.to_datetime(raw_dob, errors="coerce")
                dob_norm = dt.strftime("%m/%d/%Y") if not pd.isna(dt) else (str(raw_dob).strip() if raw_dob is not None else None)
                rec["Driver Date of Birth"] = f"{dob_norm}" if dob_norm else None
            except Exception:
                rec["Driver Date of Birth"] = f"{str(raw_dob).strip()}" if raw_dob not in (None, "") else None
            rec["_full_name_std"] = self._std_upper(rec.get("Driver Full Name")) if rec.get("Driver Full Name") else ""
            rec["_dob_iso"] = self._dob_iso(rec.get("Driver Date of Birth"))
            rec["CDL Number"] = key
            rec["CDL Type"] = group[cdl_type_col].dropna().iloc[0] if (cdl_type_col in group and not group[cdl_type_col].dropna().empty) else None
            rec["Lic State"] = group[lic_state_col].dropna().iloc[0] if (lic_state_col in group and not group[lic_state_col].dropna().empty) else None
            rec["Passenger Endt"] = group[pass_end_col].dropna().iloc[0] if (pass_end_col in group and not group[pass_end_col].dropna().empty) else None
            rec["School Bus Endt"] = group[school_end_col].dropna().iloc[0] if (school_end_col in group and not group[school_end_col].dropna().empty) else None
            rec["LIC Status"] = group[lic_status_col].dropna().astype(str).iloc[0] if (lic_status_col in group and not group[lic_status_col].dropna().empty) else None
            notes = []
            if viol_desc_col in group:
                notes = [str(x).strip() for x in group[viol_desc_col].dropna().astype(str).unique()]
            rec["Notes"] = " ; ".join([n for n in notes if n and n.lower() != "nan"]) if notes else None
            cat_series = group[viol_cat_col].astype(str) if (viol_cat_col and viol_cat_col in group) else pd.Series([""] * len(group))
            desc_series = group[viol_desc_col].astype(str) if (viol_desc_col and viol_desc_col in group) else pd.Series([""] * len(group))
            rec["# Accidents"] = int(cat_series.str.contains("accident", case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("accident", case=False, na=False).sum())
            rec["# Minor Violations"] = int(cat_series.str.contains("minor", case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("minor", case=False, na=False).sum())
            rec["# MAJOR Violations"] = int(cat_series.str.contains("|".join(["major","dui","dwi","felony","suspension"]), case=False, na=False).sum()) if not cat_series.empty else int(desc_series.str.contains("|".join(["major","dui","dwi","felony","suspension"]), case=False, na=False).sum())
            rec["YES"] = "X" if str(rec.get("LIC Status","")).strip().upper() == "VALID" else None
            rec["NO"] = None
            records.append(rec)
        df_records = pd.DataFrame(records)

        # matching: CDL primary, Name+DOB fallback; keep track of matched lookup rows
        fallback_matches = []
        unmatched_mask_indices = []
        matched_lookup_indices = set()
        
        for idx, rec in df_records.iterrows():
            cdl_key = str(rec.get("_cdl_key", "")).strip()
            m_name_std = rec.get("_full_name_std")
            m_dob_iso = rec.get("_dob_iso")
            
            matched_by = None
            found_lookup_idx = None
            
            # --- Strategy 1: CDL Match (Exact or Cleaned) ---
            candidate_idxs = []
            temp_matched_by = None
            
            # Only attempt CDL match if it's a real CDL key (not our dummy _NO_CDL_ key)
            if cdl_key and not cdl_key.startswith("_NO_CDL_"):
                # 1a. Exact Match
                if cdl_key in lookup_key_to_indices:
                    candidate_idxs = lookup_key_to_indices.get(cdl_key, [])
                    temp_matched_by = "CDL"
                else:
                    # 1b. Cleaned Match
                    try:
                        cclean = self._clean_cdl_key(cdl_key)
                    except Exception:
                        cclean = ""
                    if cclean and cclean in lookup_key_to_indices:
                        candidate_idxs = lookup_key_to_indices.get(cclean, [])
                        temp_matched_by = "CDL (cleaned)"
            
            if candidate_idxs:
                # Resolve best match among candidates (if multiple people share CDL or collision)
                best_i = None
                best_score = -1
                for i in candidate_idxs:
                    lr = lookup_rows[i]
                    lk_name_std = lr.get("_lookup_fullname_std")
                    lk_dob_iso = lr.get("_lookup_dob_iso")
                    score = 0
                    if m_dob_iso and lk_dob_iso and m_dob_iso == lk_dob_iso:
                        score += 5
                    score += self._score_name_match(m_name_std, lk_name_std)
                    if score > best_score:
                        best_score = score
                        best_i = i
                
                if best_i is not None:
                    found_lookup_idx = best_i
                    matched_by = temp_matched_by
                elif candidate_idxs:
                    found_lookup_idx = candidate_idxs[0]
                    matched_by = temp_matched_by

            # --- Strategy 2: Name + DOB Match (Fallback) ---
            # If no CDL match found (or CDL missing), try Name + DOB
            if found_lookup_idx is None:
                for i, lr in enumerate(lookup_rows):
                    if i in matched_lookup_indices:
                        continue
                    
                    lk_name_std = lr.get("_lookup_fullname_std")
                    lk_dob = lr.get("_lookup_dob")
                    
                    try:
                        # Use the robust bidirectional name matching
                        if self._date_equal(rec.get("Driver Date of Birth"), lk_dob) and self._is_name_match(m_name_std, lk_name_std):
                            found_lookup_idx = i
                            matched_by = "Name+DOB"
                            break
                    except Exception as e:
                        # self.logger.debug("Error during flexible name+DOB match: %s", e, exc_info=True)
                        continue

            # --- Apply Match Data ---
            doh_raw = None
            if found_lookup_idx is not None:
                matched_lookup_indices.add(found_lookup_idx)
                lr = lookup_rows[found_lookup_idx]
                doh_raw = lr.get(hire_col_lookup)
                
                # Update DOB if missing in MVR
                lookup_dob = lr.get("_lookup_dob")
                cur_dob = rec.get("Driver Date of Birth")
                if (cur_dob is None or (isinstance(cur_dob, float) and pd.isna(cur_dob)) or (isinstance(cur_dob, str) and str(cur_dob).strip() == "")):
                    if lookup_dob not in (None, ""):
                        df_records.at[idx, "Driver Date of Birth"] = str(lookup_dob)
                
                if matched_by == "Name+DOB":
                     fallback_matches.append({
                        "MVR_CDL": cdl_key or None,
                        "Driver": rec.get("Driver Full Name"),
                        "MVR_DOB": rec.get("Driver Date of Birth"),
                        "Matched Lookup CDL": lr.get("_cdl_key_norm"),
                        "Matched Lookup Name": lr.get("_lookup_name"),
                        "Matched Hire Raw": doh_raw
                    })
            else:
                unmatched_mask_indices.append(idx)

            df_records.at[idx, "DOH_raw"] = doh_raw
            df_records.at[idx, "DOH"] = self._format_doh_for_excel(doh_raw) if doh_raw is not None else None
            df_records.at[idx, "MatchedBy"] = matched_by

        # append lookup-only rows (those not matched)
        appended_lookup_rows = []
        for i, lr in enumerate(lookup_rows):
            if i in matched_lookup_indices:
                continue
            append_rec = {}
            append_rec["_cdl_key"] = lr.get("_cdl_key_norm") or None
            name_val = lr.get("_lookup_name") or lr.get(next((c for c in lookup_df.columns if "name" in str(c).lower()), None), None)
            append_rec["Driver Full Name"] = name_val
            dob_val = lr.get("_lookup_dob") or None
            append_rec["Driver Date of Birth"] = f"{dob_val}" if dob_val not in (None, "") else None
            append_rec["CDL Number"] = lr.get("_cdl_key_norm")
            append_rec["CDL Type"] = None
            append_rec["Lic State"] = None
            append_rec["Passenger Endt"] = None
            append_rec["School Bus Endt"] = None
            append_rec["LIC Status"] = None
            append_rec["Notes"] = "Missing MVR"
            append_rec["# Accidents"] = 0
            append_rec["# Minor Violations"] = 0
            append_rec["# MAJOR Violations"] = 0
            append_rec["YES"] = None
            append_rec["NO"] = None
            doh_raw = lr.get(hire_col_lookup)
            append_rec["DOH_raw"] = doh_raw
            append_rec["DOH"] = self._format_doh_for_excel(doh_raw) if doh_raw is not None else None
            append_rec["MatchedBy"] = "LookupOnly"
            appended_lookup_rows.append(append_rec)
        if appended_lookup_rows:
            df_appends = pd.DataFrame(appended_lookup_rows)
            df_records = pd.concat([df_records, df_appends], ignore_index=True)

        # compute Age from DOB (strip apostrophe)
        df_records["Age"] = df_records["Driver Date of Birth"].apply(self._compute_age_from_str_dob)

        return df_records

    def generate_report(self, df_records: pd.DataFrame, main_bytes: bytes):
        # load main workbook (openpyxl) for replication
        main_wb = load_workbook(io.BytesIO(main_bytes), data_only=False)
        if self.MVR_PREFERRED_NAME in main_wb.sheetnames:
            src_mvr_ws = main_wb[self.MVR_PREFERRED_NAME]
        else:
            src_mvr_ws = main_wb[main_wb.sheetnames[0]]

        # write to template
        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found at {self.TEMPLATE_PATH}")
        wb = load_workbook(self.TEMPLATE_PATH)
        if self.ALLTRANS_SHEET not in wb.sheetnames:
            raise ValueError(f"Template must contain sheet '{self.ALLTRANS_SHEET}'")
        ws = wb[self.ALLTRANS_SHEET]
        header_cells = list(ws[self.ALLTRANS_HEADER_ROW])
        template_headers = [c.value for c in header_cells]
        data_start_row = self.ALLTRANS_HEADER_ROW + 1

        for idx, rec in df_records.reset_index(drop=True).iterrows():
            write_row = data_start_row + idx
            for col_idx, hdr in enumerate(template_headers, start=1):
                if hdr is None:
                    continue
                h = str(hdr).strip().lower()
                val = None
                if h in ("name of driver","driver full name","name","driver"):
                    val = rec.get("Driver Full Name")
                elif h in ("dob","date of birth"):
                    val = rec.get("Driver Date of Birth")
                elif h == "age":
                    val = rec.get("Age")
                elif h in ("lic state","license state","state"):
                    val = rec.get("Lic State")
                elif h in ("lic class","license class","cdl type","class"):
                    val = rec.get("CDL Type")
                elif "passenger" in h:
                    val = rec.get("Passenger Endt")
                elif "school" in h:
                    val = rec.get("School Bus Endt")
                elif "accident" in h:
                    val = rec.get("# Accidents")
                elif "minor" in h:
                    val = rec.get("# Minor Violations")
                elif "major" in h:
                    val = rec.get("# MAJOR Violations")
                elif h.upper() == "YES":
                    val = rec.get("YES")
                elif h.upper() == "NO":
                    val = None
                elif h in ("notes",):
                    existing_notes = rec.get("Notes") or ""
                    if rec.get("MatchedBy") == "LookupOnly":
                        if existing_notes:
                            val = f"{existing_notes} Missing MVR"
                        else:
                            val = "Missing MVR"
                    else:
                        val = existing_notes if existing_notes else None
                elif h in ("doh","date of hire","hire date"):
                    val = rec.get("DOH")
                elif "af claims" in h:
                    val = None
                elif "lic status" in h or "license status" in h:
                    val = rec.get("LIC Status")
                else:
                    if hdr in df_records.columns:
                        val = rec.get(hdr)
                try:
                    ws.cell(row=write_row, column=col_idx, value=val)
                except Exception as e:
                    self.logger.debug("Failed writing cell r=%s c=%s val=%s: %s", write_row, col_idx, val, e, exc_info=True)

        # replicate MVR sheet exactly
        self._replicate_sheet_across_workbooks(src_mvr_ws, wb, "MVR")

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return out

    def run(self, main_bytes: bytes, lookup_bytes: bytes, chosen_lookup_sheet: str = None, preview_rows: int = 8):
        df_records = self.process_data(main_bytes, lookup_bytes, chosen_lookup_sheet, preview_rows)
        return self.generate_report(df_records, main_bytes)
